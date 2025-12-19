import json
import requests
import pandas as pd
from typing import List, Generator, Dict, Any, Literal, Optional
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
try:
    from tqdm.auto import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

from .exceptions import (
    UnihraError, UnihraApiError, UnihraConnectionError, 
    UnihraValidationError, UnihraDependencyError, raise_for_error_code
)

BASE_URL = "https://unihra.ru"
ACTION_MAP = {
    "Добавить": "add",
    "Увеличить": "increase",
    "Уменьшить": "decrease",
    "Ок": "ok",
    "Ничего не делать": "ok"
}

class UnihraClient:
    """
    Official Python Client for Unihra API.
    
    Features:
    - Automatic SSE stream handling.
    - Response normalization (converts API keys to snake_case).
    - Smart retries for network stability.
    - Pandas and Excel export integration (Multi-sheet support).
    - Visual progress bars for Jupyter/Console.
    """

    def __init__(self, api_key: str, base_url: str = BASE_URL, max_retries: int = 0):
        """
        Initialize the client.

        :param api_key: Your API key.
        :param base_url: Base URL for the API.
        :param max_retries: Number of retries for failed requests (429/50x). 
                            Default is 0 (fail fast). Set to 3-5 for production.
        """
        self.base_url = base_url.rstrip('/')
        self.api_v1 = f"{self.base_url}/api/v1"
        self.session = requests.Session()
        
        # Standard headers
        self.session.headers.update({
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "UnihraPythonSDK/1.3.0"
        })

        # Configure Smart Retries (Exponential Backoff)
        if max_retries > 0:
            retry_strategy = Retry(
                total=max_retries,
                backoff_factor=2,  # Wait 2s, 4s, 8s...
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["POST", "GET"]
            )
            adapter = HTTPAdapter(max_retries=retry_strategy)
            self.session.mount("https://", adapter)
            self.session.mount("http://", adapter)

    def health(self) -> Dict[str, Any]:
        """Check API service availability."""
        try:
            resp = self.session.get(f"{self.api_v1}/health")
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"Health check failed: {e}")

    def analyze(
        self, 
        own_page: str, 
        competitors: List[str],
        queries: Optional[List[str]] = None,
        lang: Literal['ru', 'en'] = 'ru',
        verbose: bool = False
    ) -> Dict[str, Any]:
        """
        Run a full SEO analysis (Synchronous).
        Blocks execution until the task is complete.

        :param own_page: URL of the target page.
        :param competitors: List of competitor URLs.
        :param queries: List of target search queries for context analysis.
        :param lang: Language code ('ru' or 'en').
        :param verbose: If True, displays a progress bar (requires 'tqdm').
        :return: Dictionary containing the analysis result with normalized keys.
        """
        last_event = {}
        
        # Setup Progress Bar
        pbar = None
        if verbose:
            if TQDM_AVAILABLE:
                # 0 to 100%
                pbar = tqdm(total=100, desc="Analyzing SEO", unit="%")
            else:
                print("Note: Install 'tqdm' to see a visual progress bar.")

        try:
            for event in self.analyze_stream(own_page, competitors, queries, lang):
                last_event = event
                
                # Update Progress Bar
                if pbar:
                    state = event.get("state")
                    progress = event.get("progress", 0)
                    
                    if isinstance(progress, (int, float)):
                        pbar.n = int(progress)
                        pbar.refresh()
                    
                    if state == "PROCESSING" or state == "PROGRESS":
                        msg = "Processing"
                        # Try to get detailed message
                        details = event.get("details", {})
                        if isinstance(details, dict) and "message" in details:
                            msg = details["message"][:40] # Truncate for display
                        pbar.set_description(f"{msg}")
                    elif state == "SUCCESS":
                        pbar.set_description("Completed ✅")
                        pbar.n = 100
                        pbar.refresh()

                if event.get("state") == "SUCCESS":
                    return event.get("result", {})
                    
        except Exception as e:
            if pbar: 
                pbar.set_description("Failed ❌")
                pbar.close()
            raise e
        finally:
            if pbar: 
                pbar.close()
        
        return last_event

    def analyze_stream(
        self, 
        own_page: str, 
        competitors: List[str],
        queries: Optional[List[str]] = None,
        lang: str = 'ru'
    ) -> Generator[Dict, None, None]:
        """
        Generator method for real-time updates.
        Yields SSE events as dictionaries.
        
        Automatically normalizes API keys (e.g. 'Block Comparison' -> 'block_comparison').
        """
        # 1. Validation
        if not competitors:
            raise UnihraValidationError("Competitor list cannot be empty.")

        payload = {
            "own_page": own_page, 
            "competitor_urls": competitors,
            "queries": queries or [],
            "lang": lang
        }

        try:
            # 2. Create Task
            resp = self.session.post(f"{self.api_v1}/process", json=payload)
            
            if resp.status_code == 401:
                raise UnihraApiError("Invalid API Key or unauthorized access", code=401)
            resp.raise_for_status()
            
            task_id = resp.json().get("task_id")
            if not task_id:
                raise UnihraApiError("API response missing 'task_id'")

            # 3. Stream Results
            stream_url = f"{self.api_v1}/process/status/{task_id}"
            
            with self.session.get(stream_url, stream=True) as s_resp:
                s_resp.raise_for_status()
                
                for line in s_resp.iter_lines():
                    if not line: 
                        continue
                    
                    if line.startswith(b'data: '):
                        try:
                            # Decode SSE JSON
                            decoded_line = line[6:].decode('utf-8')
                            data = json.loads(decoded_line)
                            state = data.get("state")
                            
                            # --- ERROR HANDLING ---
                            if state == "FAILURE":
                                # Handle nested error object: {"error": {"code": 1003}}
                                error_obj = data.get("error")
                                if isinstance(error_obj, dict):
                                    code = error_obj.get("code", 9999)
                                    msg = error_obj.get("message", "Unknown error")
                                else:
                                    # Handle flat structure
                                    code = data.get("error_code", 9999)
                                    msg = data.get("message", "Unknown error")

                                raise_for_error_code(code, msg, data)
                            
                            if state == "SUCCESS":
                                raw_result = data.get("result", {})
                                normalized_result = self._normalize_keys(raw_result)
                                if lang == 'en':
                                    final_result = self._translate_action_values(normalized_result)
                                else:
                                    final_result = normalized_result

                                data["result"] = final_result
                                yield data
                                break
                            
                            yield data
                                
                        except json.JSONDecodeError:
                            continue
                            
        except requests.exceptions.RetryError:
            raise UnihraConnectionError("Max retries exceeded. The service might be temporarily unavailable.")
        except requests.exceptions.RequestException as e:
            raise UnihraConnectionError(f"Network error: {e}")

    def _normalize_keys(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Helper to convert API keys to Pythonic snake_case.
        Example: 'Block Comparison' -> 'block_comparison'
        """
        new_data = {}
        for key, value in data.items():
            new_key = key.lower().replace(" ", "_").replace("-", "_")
            new_data[new_key] = value
        return new_data

    def _translate_action_values(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Translates 'action_needed' values from Russian to English."""
        if "block_comparison" in result and isinstance(result["block_comparison"], list):
            for item in result["block_comparison"]:
                if "action_needed" in item:
                    russian_action = item["action_needed"]
                    # Use .get() to safely handle unknown values from API
                    item["action_needed"] = ACTION_MAP.get(russian_action, russian_action)
        return result

    def get_dataframe(self, result: Dict[str, Any], section: str = "block_comparison") -> pd.DataFrame:
        """
        Convert a specific result section to a Pandas DataFrame.
        
        :param result: The dictionary returned by .analyze()
        :param section: 'block_comparison', 'ngrams_analysis', etc.
        :return: pandas.DataFrame
        """
        try:
            import pandas as pd
        except ImportError:
            raise UnihraDependencyError("Pandas is not installed. Run: pip install pandas")

        # Normalize section name just in case user passes "Block Comparison"
        normalized_section = section.lower().replace(" ", "_").replace("-", "_")
        data = result.get(normalized_section, [])
        return pd.DataFrame(data)

    def save_report(self, result: Dict[str, Any], filename: str = "report.xlsx", style_output: bool = True):
        """
        Save the full analysis result to a file.
        Includes Semantic Gaps, Block Comparison, N-Grams, and DrMaxs.
        """
        try:
            import pandas as pd
        except ImportError:
            raise UnihraDependencyError("Pandas is required. Run: pip install pandas openpyxl")

        df_blocks = pd.DataFrame(result.get("block_comparison", []))
        
        # Handle N-grams normalization
        ngrams_data = result.get("ngrams_analysis") or result.get("n_grams_analysis") or []
        df_ngrams = pd.DataFrame(ngrams_data)
        
        # New Semantic Context (Gaps) Data
        # Can be either 'semantic_context_analysis' or legacy 'semantic_context_gaps'
        gaps_data = result.get("semantic_context_analysis") or result.get("semantic_context_gaps") or []
        df_gaps = pd.DataFrame(gaps_data)
        
        drmaxs_data = result.get("drmaxs", {})

        if filename.endswith(".csv"):
            # CSV export is limited to the main block comparison for simplicity
            df_blocks.to_csv(filename, index=False, encoding='utf-8-sig')
        else:
            try:
                import openpyxl
            except ImportError:
                raise UnihraDependencyError("Library 'openpyxl' is required for Excel export.")

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # 1. Semantic Gaps (High Priority)
                if not df_gaps.empty:
                    sheet = "Semantic Gaps"
                    # Reorder columns for readability if they exist in the DataFrame
                    desired_cols = ['lemma', 'recommendation', 'context_snippet', 'gap', 'coverage_percent', 'competitor_avg_score', 'own_score']
                    existing_cols = [c for c in desired_cols if c in df_gaps.columns]
                    other_cols = [c for c in df_gaps.columns if c not in desired_cols]
                    
                    df_gaps_ordered = df_gaps[existing_cols + other_cols]
                    df_gaps_ordered.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_gaps_ordered, sheet_type="gaps")

                # 2. Word Analysis
                if not df_blocks.empty:
                    sheet = "Word Analysis"
                    df_blocks.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_blocks, sheet_type="word_analysis")
                
                # 3. N-Grams
                if not df_ngrams.empty:
                    sheet = "N-Grams"
                    df_ngrams.to_excel(writer, sheet_name=sheet, index=False)
                    if style_output: self._style_worksheet(writer.sheets[sheet], df_ngrams, sheet_type="ngrams")
                
                # 4. DrMaxs Vectors
                if drmaxs_data and isinstance(drmaxs_data, dict):
                    for subkey, subdata in drmaxs_data.items():
                        if subdata and isinstance(subdata, list):
                            df_dr = pd.DataFrame(subdata)
                            safe_name = subkey.replace("_", " ").title().replace("By", "")
                            sheet_name = f"Vectors {safe_name}"[:31] # Excel limit is 31 chars
                            df_dr.to_excel(writer, sheet_name=sheet_name, index=False)
                            if style_output: self._style_worksheet(writer.sheets[sheet_name], df_dr, sheet_type="vectors")

    def _style_worksheet(self, worksheet, df, sheet_type="generic"):
        """
        Internal method to apply professional styling:
        1. Auto-width for columns.
        2. Conditional formatting based on sheet type and values.
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="363636", end_color="363636", fill_type="solid")
        
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 1. Format Headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 2. Auto-width
        for idx, col in enumerate(df.columns):
            # Calculate max length of content or header
            max_len = max(
                [len(str(s)) for s in df[col].astype(str).values] + [len(col)]
            )
            # Cap width
            final_width = min(max_len + 2, 60) 
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = final_width

        # 3. Conditional Formatting
        
        # Helper to map dataframe column names to Excel column indices (1-based)
        col_map = {name: i + 1 for i, name in enumerate(df.columns)}

        if sheet_type == "gaps":
            # In Gaps list: if own_score = 0, color 'lemma' cells. 
            # Assuming 0 is bad (Red), >0 is good (Green).
            if 'own_score' in col_map and 'lemma' in col_map:
                score_idx = col_map['own_score']
                lemma_idx = col_map['lemma']
                
                for row in range(2, worksheet.max_row + 1):
                    score_val = worksheet.cell(row=row, column=score_idx).value
                    # Check if score is effectively 0
                    try:
                        is_missing = float(score_val) == 0 if score_val is not None else True
                    except (ValueError, TypeError):
                        is_missing = True
                    
                    if is_missing:
                        worksheet.cell(row=row, column=lemma_idx).fill = red_fill
                    else:
                        worksheet.cell(row=row, column=lemma_idx).fill = green_fill

        else:
            # For Word Analysis, N-Grams, and Vectors
            # Determine target columns based on sheet type
            target_cols = []
            if sheet_type == "word_analysis":
                # Paint id, block_id, word, lemma
                target_names = ["id", "block_id", "word", "lemma"]
                target_cols = [col_map[c] for c in target_names if c in col_map]
            elif sheet_type == "ngrams":
                # Paint id, block_id, ngram
                target_names = ["id", "block_id", "ngram"]
                target_cols = [col_map[c] for c in target_names if c in col_map]
            elif sheet_type == "vectors":
                # Paint id, analysis_id, word
                target_names = ["id", "analysis_id", "word"]
                target_cols = [col_map[c] for c in target_names if c in col_map]

            # Logic: Check bool 'present_on_own_page'
            bool_col = 'present_on_own_page'
            # Fallback for vectors if column name differs but logic implies existence check
            if bool_col not in col_map and 'present_in_own' in col_map:
                 bool_col = 'present_in_own'

            if bool_col in col_map and target_cols:
                bool_idx = col_map[bool_col]
                
                for row in range(2, worksheet.max_row + 1):
                    is_present = worksheet.cell(row=row, column=bool_idx).value
                    
                    fill_color = None
                    if is_present is True:
                        fill_color = green_fill
                    elif is_present is False:
                        fill_color = red_fill
                    
                    if fill_color:
                        for t_idx in target_cols:
                            worksheet.cell(row=row, column=t_idx).fill = fill_color