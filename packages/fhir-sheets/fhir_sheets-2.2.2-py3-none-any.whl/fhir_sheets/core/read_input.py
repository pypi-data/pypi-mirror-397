from typing import List
import openpyxl
import logging

from .model.cohort_data_entity import CohortData, CohortData

from .model.resource_definition_entity import ResourceDefinition
from .model.resource_link_entity import ResourceLink

logger: logging.Logger = logging.getLogger("fhirsheets.core.read_input")

# Function to read the xlsx file and access specific sheets
def read_xlsx_and_process(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    resource_definition_entities = []
    resource_link_entities = []
    cohort_data = CohortData.from_dict([],[])
    # Example of accessing specific sheets
    if 'ResourceDefinitions' in workbook.sheetnames:
        sheet = workbook['ResourceDefinitions']
        resource_definition_entities = process_sheet_resource_definitions(sheet)

    if 'ResourceLinks' in workbook.sheetnames:
        sheet = workbook['ResourceLinks']
        resource_link_entities = process_sheet_resource_links(sheet)

    if 'PatientData' in workbook.sheetnames:
        sheet = workbook['PatientData']
        cohort_data = process_sheet_patient_data_revised(sheet, resource_definition_entities)
    
    return resource_definition_entities, resource_link_entities, cohort_data


# Function to process the specific sheet with 'Entity Name', 'ResourceType', and 'Profile(s)'
def process_sheet_resource_definitions(sheet) -> List[ResourceDefinition]:
    resource_definitions = []
    resource_definition_entities = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]  # Get headers

    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = dict((h, r) for h, r in zip(headers, row) if h is not None)  # Create a dictionary for each row
        if all(cell is None or cell == "" for cell in row_data.values()):
            continue
        # Split 'Profile(s)' column into a list of URLs
        if row_data.get("Profile(s)"):
            row_data["Profile(s)"] = [url.strip() for url in row_data["Profile(s)"].split(",")]
        resource_definition_entities.append(ResourceDefinition.from_dict(row_data))
        resource_definitions.append(row_data)
    logger.info(f"Resource Definitions\n----------{resource_definitions}")
    return resource_definition_entities

# Function to process the specific sheet with 'OriginResource', 'ReferencePath', and 'DestinationResource'
def process_sheet_resource_links(sheet) -> List[ResourceLink]:
    resource_links = []
    resource_link_entities = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]  # Get headers
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = dict(zip(headers, row))  # Create a dictionary for each row
        if all(cell is None or cell == "" for cell in row_data):
            continue
        resource_links.append(row_data)
        resource_link_entities.append(ResourceLink.from_dict(row_data))
    logger.info(f"Resource Links\n----------{resource_links}")
    return resource_link_entities

# Function to process the "PatientData" sheet for the Revised CohortData
def process_sheet_patient_data_revised(sheet, resource_definition_entities):
    headers = []
    patients = []
    # Initialize the dictionary to store the processed data
    # Process the Header Entries from the first 6 rows (Entity To Query, JsonPath, etc.) and the data from the rest.
    for col in sheet.iter_cols(min_row=1, min_col=3, values_only=True):  # Start from 3rd column
        if all(entry is None for entry in col):
            continue
        entity_name = col[0]  # The entity name comes from the first row (Entity To Query)
        field_name = col[5]  #The "Data Element" comes from the fifth row
        if (entity_name is None or entity_name == "") and (field_name is not None and field_name != ""):
            logger.warning(f"Reading Patient Data Issue - {field_name} - 'Entity To Query' cell missing for column labelled '{field_name}', please provide entity name from the ResourceDefinitions tab.")

        if entity_name not in [entry.entityName for entry in resource_definition_entities]:
            logger.warning(f"Reading Patient Data Issue - {field_name} - 'Entity To Query' cell has entity named '{entity_name}', however, the ResourceDefinition tab has no matching resource. Please provide a corresponding entry in the ResourceDefinition tab.")

        # Create a header entry
        header_data = {
            "fieldName": field_name,
            "entityName": entity_name,
            "jsonPath": col[1],  # JsonPath from the second row
            "valueType": col[2], # Value Type from the third row
            "valueSets": col[3] # Value Set from the fourth row
        }
        headers.append(header_data)
        # Create a data entry
        values = col[6:] # The values come from the 6th row and below
        values = tuple(item for item in values if item is not None)
        #Expand the patient dictionary set if needed
        if len(values) > len(patients):
            needed_count = len(values) - len(patients)
            patients.extend([{}] * needed_count)
        for patient_dict, value in zip(patients, values):
            patient_dict[(entity_name, field_name)] = value
    logger.info(f"Headers\n----------{headers}")
    logger.info(f"Patients\n----------{patients}")
    cohort_data = CohortData.from_dict(headers=headers, patients=patients)
    return cohort_data