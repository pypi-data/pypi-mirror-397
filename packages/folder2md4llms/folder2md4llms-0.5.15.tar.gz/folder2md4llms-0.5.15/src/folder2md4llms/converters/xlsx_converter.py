"""XLSX document converter."""

import logging
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook

    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

from .base import BaseConverter, ConversionError

logger = logging.getLogger(__name__)


class XLSXConverter(BaseConverter):
    """Converts XLSX files to text."""

    def __init__(self, config: dict[str, Any] | None = None):
        super().__init__(config)
        self.max_sheets = self.config.get("xlsx_max_sheets", 10)
        self.max_rows = self.config.get("xlsx_max_rows", 100)
        self.max_cols = self.config.get("xlsx_max_cols", 26)

    def can_convert(self, file_path: Path) -> bool:
        """Check if this converter can handle the given file."""
        return (
            XLSX_AVAILABLE
            and file_path.suffix.lower() in {".xlsx", ".xls"}
            and file_path.exists()
        )

    def convert(self, file_path: Path) -> str | None:
        """Convert XLSX to text."""
        if not XLSX_AVAILABLE:
            return (
                "XLSX conversion not available. Install openpyxl: pip install openpyxl"
            )

        try:
            workbook = load_workbook(file_path, read_only=True)

            text_parts = []
            text_parts.append(f"Excel Workbook: {file_path.name}")
            text_parts.append(f"Sheets: {', '.join(workbook.sheetnames)}")
            text_parts.append("=" * 50)
            text_parts.append("")

            # Process sheets
            sheets_processed = 0
            for sheet_name in workbook.sheetnames:
                if sheets_processed >= self.max_sheets:
                    text_parts.append(f"... (showing first {self.max_sheets} sheets)")
                    break

                sheet = workbook[sheet_name]
                text_parts.append(f"Sheet: {sheet_name}")
                text_parts.append("-" * 30)
                text_parts.append("")

                # Get sheet data
                sheet_data = self._extract_sheet_data(sheet)

                if sheet_data:
                    # Convert to markdown table
                    text_parts.append(self._format_as_markdown_table(sheet_data))
                else:
                    text_parts.append("(Empty sheet)")

                text_parts.append("")
                text_parts.append("")
                sheets_processed += 1

            return "\n".join(text_parts)

        except Exception as e:
            logger.error(f"Error converting XLSX {file_path}: {e}")
            raise ConversionError(f"Failed to convert XLSX: {str(e)}") from e

    def _extract_sheet_data(self, sheet) -> list:
        """Extract data from a worksheet."""
        data = []

        try:
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= self.max_rows:
                    break

                # Skip completely empty rows
                if not any(cell for cell in row):
                    continue

                # Limit columns
                limited_row = row[: self.max_cols]

                # Convert None values to empty strings
                cleaned_row = [
                    str(cell) if cell is not None else "" for cell in limited_row
                ]
                data.append(cleaned_row)
                row_count += 1

            return data

        except Exception as e:
            logger.error(f"Error extracting sheet data: {e}")
            return []

    def _format_as_markdown_table(self, data: list) -> str:
        """Format data as a markdown table."""
        if not data:
            return ""

        # Find the maximum number of columns
        max_cols = max(len(row) for row in data) if data else 0

        # Pad rows to have the same number of columns
        padded_data = []
        for row in data:
            padded_row = row + [""] * (max_cols - len(row))
            padded_data.append(padded_row)

        # Create markdown table
        lines = []

        if padded_data:
            # Header row (first row)
            header = padded_data[0]
            lines.append("| " + " | ".join(header) + " |")

            # Separator row
            lines.append(
                "| " + " | ".join(["-" * max(1, len(cell)) for cell in header]) + " |"
            )

            # Data rows
            for row in padded_data[1:]:
                lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines)

    def get_supported_extensions(self) -> set:
        """Get the file extensions this converter supports."""
        return {".xlsx", ".xls"}

    def get_document_info(self, file_path: Path) -> dict[str, Any]:
        """Get XLSX-specific information."""
        info = self.get_file_info(file_path)

        if not XLSX_AVAILABLE:
            info["error"] = "XLSX library not available"
            return info

        try:
            workbook = load_workbook(file_path, read_only=True)

            info.update(
                {
                    "sheets": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                }
            )

            # Get row/column counts for each sheet
            sheet_info = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info[sheet_name] = {
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }

            info["sheet_details"] = sheet_info

        except Exception as e:
            info["error"] = str(e)

        return info
