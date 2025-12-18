import binascii
import logging
from typing import AsyncGenerator

import xlrd
from openpyxl.reader.excel import load_workbook

from docsloader import utils
from docsloader.base import BaseLoader, DocsData

logger = logging.getLogger(__name__)


class XlsxLoader(BaseLoader):
    """
    xlsx loader

    params:
        - path_or_url: str
        - suffix: str = None
        - encoding: str = None
        - load_type: str = "basic"
        - load_options: dict = None
        - metadata: dict = None
        - rm_tmpfile: bool = False
    """

    async def load_by_basic(self) -> AsyncGenerator[DocsData, None]:
        table_fmt = self.load_options.get("table_fmt")
        with open(self.tmpfile, "rb") as f:
            header_flag = binascii.hexlify(f.read(8)).decode().upper()
        if header_flag.startswith("504B0304"):  # .xlsx
            wb = load_workbook(filename=self.tmpfile, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                rows = ws.iter_rows(values_only=True)
                try:
                    header = list(next(rows))
                except StopIteration:
                    header = []
                    rows = []
                self.metadata.update({
                    "header": header,
                    "sheet_name": sheet_name,
                })
                has_value = False
                for row in rows:
                    has_value = True
                    row = list(row)
                    yield DocsData(
                        type="text",
                        text=utils.format_table(row),
                        data=row,
                        metadata=self.metadata,
                    )
                if not has_value:
                    yield DocsData(
                        type="text",
                        text="",
                        data=[],
                        metadata=self.metadata,
                    )
            wb.close()
        elif header_flag.startswith("D0CF11E0A1B11AE1"):  # .xls
            book = xlrd.open_workbook(self.tmpfile, formatting_info=False)
            for sheet_name in book.sheet_names():
                sheet = book.sheet_by_name(sheet_name)
                logger.info(f"Processing sheet: {sheet_name}")
                header = sheet.row_values(0) if sheet.nrows > 0 else []
                self.metadata.update({
                    "header": header,
                    "sheet_name": sheet_name,
                })
                if sheet.nrows > 1:
                    for idx in range(1, sheet.nrows):
                        row = sheet.row_values(idx)
                        yield DocsData(
                            type="text",
                            text=utils.format_table(row, fmt=table_fmt),
                            data=row,
                            metadata=self.metadata,
                        )
                else:
                    yield DocsData(
                        type="text",
                        text="",
                        data=[],
                        metadata=self.metadata,
                    )
        else:
            raise ValueError(f"Unsupported file header: {header_flag}")
