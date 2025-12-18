#!/usr/bin/env python3
"""
jBOM - KiCad Bill of Materials Generator

Takes a KiCad project and inventory file (CSV/Excel/Numbers) to generate a bill of materials.
Matches components to inventory entries based on type, value, and attributes.
"""

import re
import sys
import csv
import argparse
import warnings
import json
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Union
from dataclasses import dataclass, field
from sexpdata import Symbol

# Suppress specific Numbers version warning
warnings.filterwarnings("ignore", message="Numbers version 14.3 not tested with this version")

# Optional imports for spreadsheet support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

try:
    from numbers_parser import Document as NumbersDocument
    NUMBERS_SUPPORT = True
except ImportError:
    NUMBERS_SUPPORT = False

# Default priority value (must be defined before classes that use it)
DEFAULT_PRIORITY = 99


def normalize_field_name(field: str) -> str:
    """
    Normalize field names to canonical snake_case format.
    Accepts: snake_case, Title Case, CamelCase, spaces, mixed formats.
    Examples: 'match_quality', 'Match Quality', 'MatchQuality', 'MATCH_QUALITY' -> 'match_quality'
    """
    if not field:
        return ''
    
    # Handle prefixes (I: and C:) separately
    prefix = ''
    if field.lower().startswith('i:'):
        prefix = 'i:'
        field = field[2:]
    elif field.lower().startswith('c:'):
        prefix = 'c:'
        field = field[2:]
    
    # Replace spaces and hyphens with underscores
    field = field.replace(' ', '_').replace('-', '_')
    
    # Insert underscores before uppercase letters (for CamelCase like MatchQuality -> match_quality)
    # But avoid double underscores
    result = []
    for i, char in enumerate(field):
        if i > 0 and char.isupper() and field[i-1].islower():
            result.append('_')
        result.append(char.lower())
    
    # Clean up multiple underscores
    normalized = ''.join(result)
    while '__' in normalized:
        normalized = normalized.replace('__', '_')
    
    return prefix + normalized.strip('_')


# Known acronyms in PCB/electronics domain (lowercase for matching)
KNOWN_ACRONYMS = {
    'lcsc', 'smd', 'pcb', 'bom', 'dnp', 'pth', 'ipn', 'mfgpn',
    'jlc', 'jlcpcb', 'csv', 'xlsx', 'usb', 'uart', 'i2c', 'spi',
    'led', 'pwm', 'adc', 'dac', 'ic', 'rf', 'esd', 'emi', 'emc',
    'url', 'pdf', 'png', 'jpg', 'via', 'gnd', 'vcc', 'can', 'psu'
}

def field_to_header(field: str) -> str:
    """
    Convert normalized field name to human-readable header for CSV.
    Uses Title Case with special handling for known acronyms.
    Examples: 
        'match_quality' -> 'Match Quality'
        'lcsc' -> 'LCSC'
        'i:package' -> 'I:Package'
        'mfgpn' -> 'MFGPN'
    """
    if not field:
        return ''
    
    # Handle prefixes
    prefix = ''
    if field.lower().startswith('i:'):
        prefix = 'I:'
        field = field[2:]
    elif field.lower().startswith('c:'):
        prefix = 'C:'
        field = field[2:]
    
    # Split on underscores and handle each part
    parts = field.split('_')
    result_parts = []
    
    for part in parts:
        if not part:
            continue
        lower_part = part.lower()
        # Check if this part is a known acronym
        if lower_part in KNOWN_ACRONYMS:
            result_parts.append(part.upper())
        else:
            result_parts.append(part.capitalize())
    
    header_part = ' '.join(result_parts)
    return prefix + header_part if prefix else header_part

@dataclass
class Component:
    """Represents a component from KiCad schematic"""
    reference: str
    lib_id: str
    value: str
    footprint: str
    properties: Dict[str, str] = field(default_factory=dict)
    in_bom: bool = True
    exclude_from_sim: bool = False
    dnp: bool = False


@dataclass
class InventoryItem:
    """Represents an item from the inventory CSV"""
    ipn: str
    keywords: str
    category: str
    description: str
    smd: str
    value: str
    type: str
    tolerance: str
    voltage: str
    amperage: str
    wattage: str
    lcsc: str
    manufacturer: str
    mfgpn: str
    datasheet: str
    package: str = ""
    priority: int = DEFAULT_PRIORITY  # Priority from CSV: 1=most desirable, higher=less desirable
    raw_data: Dict[str, str] = field(default_factory=dict)


@dataclass
class BOMEntry:
    """Represents a bill of materials entry"""
    reference: str
    quantity: int
    value: str
    footprint: str
    lcsc: str
    manufacturer: str
    mfgpn: str
    description: str
    datasheet: str
    smd: str = ""
    match_quality: str = ""
    notes: str = ""
    # Debug fields (emitted when --verbose)
    priority: int = 0


# Component type constants
class ComponentType:
    """Component type constants for standardized type identification"""
    RESISTOR = 'RES'
    CAPACITOR = 'CAP'
    INDUCTOR = 'IND'
    DIODE = 'DIO'
    LED = 'LED'
    INTEGRATED_CIRCUIT = 'IC'
    MICROCONTROLLER = 'MCU'
    TRANSISTOR = 'Q'
    CONNECTOR = 'CON'
    SWITCH = 'SWI'
    RELAY = 'RLY'
    REGULATOR = 'REG'
    OSCILLATOR = 'OSC'
    ANALOG = 'ANA'
    SILK_SCREEN = 'SLK'

# Diagnostic issue type constants
class DiagnosticIssue:
    """Diagnostic issue type constants"""
    TYPE_UNKNOWN = 'type_unknown'
    NO_TYPE_MATCH = 'no_type_match'
    NO_VALUE_MATCH = 'no_value_match'
    PACKAGE_MISMATCH = 'package_mismatch'
    PACKAGE_MISMATCH_GENERIC = 'package_mismatch_generic'
    NO_MATCH = 'no_match'

# Common inventory field constants
class CommonFields:
    """Common field name constants"""
    VOLTAGE = 'V'
    AMPERAGE = 'A' 
    WATTAGE = 'W'
    TOLERANCE = 'Tolerance'
    POWER = 'Power'
    TEMPERATURE_COEFFICIENT = 'Temperature Coefficient'

# Category-specific inventory field mappings for comprehensive property extraction
COMMON_FIELDS = [
    "IPN",          # User provided Inventory Part Number
    "Value",        # Value - usually ohms, farads... or part number
    "Description",  # Human readable
    "SMD",          # is this a Surface Mount part?
    "Manufacturer", # Human readable
    "MFGPN",        # Human readable
    "Package",      # EIA nomenclature
    "Symbol",       # KiCad Symbol
    "Footprint",    # KiCad Footprint
    "Datasheet",    # URL
]

DEFAULT_CATEGORY_FIELDS = COMMON_FIELDS + [CommonFields.VOLTAGE, CommonFields.AMPERAGE, CommonFields.WATTAGE, CommonFields.TOLERANCE, CommonFields.TEMPERATURE_COEFFICIENT]

# Category-specific field mappings with value interpretation semantics
# Note: "Value:X" means the component's Value field represents quantity X
CATEGORY_FIELDS = {
    ComponentType.ANALOG:             COMMON_FIELDS + [CommonFields.VOLTAGE],
    ComponentType.CAPACITOR:          COMMON_FIELDS + [CommonFields.VOLTAGE, 'Voltage', 'Type', CommonFields.TOLERANCE],  # Value:Capacitance
    ComponentType.CONNECTOR:          COMMON_FIELDS + ['Pitch'],
    ComponentType.DIODE:              COMMON_FIELDS + [CommonFields.VOLTAGE, CommonFields.AMPERAGE],
    ComponentType.LED:                COMMON_FIELDS + [CommonFields.VOLTAGE, CommonFields.AMPERAGE, 'mcd', 'Wavelength', 'Angle'],  # Value:Color
    ComponentType.RESISTOR:           COMMON_FIELDS + [CommonFields.VOLTAGE, CommonFields.WATTAGE, CommonFields.POWER, CommonFields.TOLERANCE],  # Value:Resistance
    ComponentType.INTEGRATED_CIRCUIT: COMMON_FIELDS + [CommonFields.VOLTAGE],
    ComponentType.INDUCTOR:           COMMON_FIELDS + [CommonFields.AMPERAGE, CommonFields.WATTAGE],  # Value:Inductance
    ComponentType.TRANSISTOR:         COMMON_FIELDS + [CommonFields.VOLTAGE, CommonFields.AMPERAGE, CommonFields.WATTAGE],
    ComponentType.MICROCONTROLLER:    COMMON_FIELDS + ['Family'],
    ComponentType.REGULATOR:          COMMON_FIELDS + [CommonFields.VOLTAGE, CommonFields.AMPERAGE, CommonFields.WATTAGE],
    ComponentType.OSCILLATOR:         COMMON_FIELDS + ['Frequency', 'Stability', 'Load'],
    ComponentType.SILK_SCREEN:        COMMON_FIELDS + ['Form'],
    ComponentType.RELAY:              COMMON_FIELDS + ['Form'],
    ComponentType.SWITCH:             COMMON_FIELDS + ['Form']
}

# Define how Value:X field should be interpreted for each category
VALUE_INTERPRETATION = {
    ComponentType.CAPACITOR: 'Capacitance',   # Value represents capacitance
    ComponentType.RESISTOR:  'Resistance',    # Value represents resistance  
    ComponentType.INDUCTOR:  'Inductance',    # Value represents inductance
    ComponentType.LED:       'Color',         # Value represents color/wavelength
}

# Package constants
class PackageType:
    """Package type lists for footprint and SMD identification"""
    SMD_PACKAGES = [ # SMD package list
        # Passive component packages (imperial)
        '0402', '0603', '0805', '1206', '1210',
        # Passive component packages (metric)
        '1005', '1608', '2012', '3216', '3225', '5050',
        # SOT packages 
        'sot', 'sot-23', 'sot-223', 'sot-89', 'sot-143', 'sot-323', 'sc-70',
        'sot-23-5', 'sot-23-6', 'sot-353', 'sot-363',
        # IC packages
        'soic', 'ssop', 'tssop', 'qfp', 'qfn', 'dfn', 'bga',
        'wlcsp', 'lga', 'plcc', 'pqfp', 'tqfp', 'lqfp', 'msop', 'sc70',
        # Diode packages
        'sod-123', 'sod-323', 'sod-523', 'sod-923',
        # Power packages (SMD)
        'dpak', 'd2pak'
    ]
    
    THROUGH_HOLE_PACKAGES = [ # Through-hole package list
        'dip', 'through-hole', 'axial', 'radial',
        'to-220', 'to-252', 'to-263', 'to-39', 'to-92'  # Through-hole power packages
    ]
    
# SMD field constants
class SMDType:
    """SMD type indicator constants"""
    SMD_VALUES = ['SMD', 'Y', 'YES', 'TRUE', '1']
    PTH_VALUES = ['PTH', 'THT', 'TH', 'THROUGH-HOLE', 'N', 'NO', 'FALSE', '0']
    UNKNOWN_VALUES = ['', 'UNKNOWN', 'N/A']

# Scoring constants
class ScoreWeights:
    """Scoring weight constants for inventory matching"""
    TOLERANCE_EXACT = 15
    TOLERANCE_BETTER = 12
    VOLTAGE_MATCH = 10
    CURRENT_MATCH = 10
    POWER_MATCH = 10
    LED_WAVELENGTH = 8
    LED_INTENSITY = 8
    OSC_FREQUENCY = 12
    OSC_STABILITY = 8
    LED_ANGLE = 5
    OSC_LOAD = 5
    CON_PITCH = 10
    MCU_FAMILY = 8
    GENERIC_PROPERTY = 3

# Precision threshold for resistor matching
PRECISION_THRESHOLD = 1.0

# Global type mapping for component type normalization
COMPONENT_TYPE_MAPPING = {
    'RESISTOR':        ComponentType.RESISTOR,
    'R':               ComponentType.RESISTOR,
    'CAPACITOR':       ComponentType.CAPACITOR,
    'C':               ComponentType.CAPACITOR,
    'DIODE':           ComponentType.DIODE,
    'D':               ComponentType.DIODE,
    'INDUCTOR':        ComponentType.INDUCTOR,
    'L':               ComponentType.INDUCTOR,
    'TRANSISTOR':      ComponentType.TRANSISTOR,
    'MICROCONTROLLER': ComponentType.MICROCONTROLLER,
    'REGULATOR':       ComponentType.REGULATOR,
    'OSCILLATOR':      ComponentType.OSCILLATOR,
    'SWITCH':          ComponentType.SWITCH,
    'RELAY':           ComponentType.RELAY,
    'CONNECTOR':       ComponentType.CONNECTOR,
    'ANALOG':          ComponentType.ANALOG
}

def normalize_component_type(component_type: str) -> str:
    """Normalize component type string to standard category using global mapping"""
    category = component_type.upper() if component_type else ''
    
    # Try direct lookup first, then mapped lookup
    if category in CATEGORY_FIELDS:
        return category
    elif category in COMPONENT_TYPE_MAPPING:
        return COMPONENT_TYPE_MAPPING[category]
    else:
        return category  # Return as-is if not found

def get_category_fields(component_type: str) -> List[str]:
    """Get relevant fields for a component category"""
    normalized_type = normalize_component_type(component_type)
    
    if normalized_type in CATEGORY_FIELDS:
        return CATEGORY_FIELDS[normalized_type]
    else:
        # Default to common fields plus some general ones
        return DEFAULT_CATEGORY_FIELDS

def get_value_interpretation(component_type: str) -> Optional[str]:
    """Get what the Value field represents for a given component category"""
    normalized_type = normalize_component_type(component_type)
    return VALUE_INTERPRETATION.get(normalized_type, None)


def get_component_type(lib_id: str, footprint: str) -> Optional[str]:
    """Determine component type from lib_id or footprint.
    
    This is used by InventoryMatcher to ensure consistent component type detection.
    
    Args:
        lib_id: Component library identifier (e.g., "Device:R", "SPCoast:resistor")
        footprint: PCB footprint name (e.g., "PCM_SPCoast:0603-RES")
    
    Returns:
        Component type string (RES, CAP, IND, etc.) or None if unrecognized
    """
    lib_id = lib_id.lower()
    footprint = footprint.lower()
    
    if 'resistor' in lib_id or 'r' == lib_id.split(':')[-1] or 'res' in footprint:
        return ComponentType.RESISTOR
    elif 'capacitor' in lib_id or 'c' == lib_id.split(':')[-1] or 'cap' in footprint:
        return ComponentType.CAPACITOR
    elif 'diode' in lib_id or 'd' == lib_id.split(':')[-1] or 'diode' in footprint:
        return ComponentType.DIODE
    elif 'led' in lib_id or 'led' in footprint:
        return ComponentType.LED
    elif 'inductor' in lib_id or 'l' == lib_id.split(':')[-1]:
        return ComponentType.INDUCTOR
    elif 'connector' in lib_id or 'conn' in lib_id:
        return ComponentType.CONNECTOR
    elif 'switch' in lib_id or 'sw' in lib_id:
        return ComponentType.SWITCH
    elif 'transistor' in lib_id or lib_id.split(':')[-1].startswith('q'):
        return ComponentType.TRANSISTOR
    elif ('ic' in lib_id or 'mcu' in lib_id or 'microcontroller' in lib_id or 
          lib_id.split(':')[-1].startswith('u') or lib_id.split(':')[-1] == 'ic'):
        return ComponentType.INTEGRATED_CIRCUIT
    
    return None

class KiCadParser:
    """Parser for KiCad schematic files using S-expression parser (sexpdata)."""

    def __init__(self, schematic_path: Path):
        self.schematic_path = schematic_path
        self.components: List[Component] = []

    def parse(self) -> List[Component]:
        """Parse the KiCad schematic file and extract components"""
        return self._parse_with_sexp()


    def _parse_with_sexp(self) -> List[Component]:
        from jbom.common.sexp_parser import load_kicad_file, walk_nodes

        sexp = load_kicad_file(self.schematic_path)
        for symbol_node in walk_nodes(sexp, 'symbol'):
            comp = self._parse_symbol(symbol_node)
            if comp and comp.in_bom and not comp.dnp and not comp.reference.startswith('#'):
                self.components.append(comp)
        return self.components

    def _parse_symbol(self, node: list) -> Optional[Component]:
        """Parse a (symbol ...) node into a Component"""
        lib_id = ""
        reference = ""
        value = ""
        footprint = ""
        in_bom = True
        exclude_from_sim = False
        dnp = False
        properties: Dict[str, str] = {}

        # Iterate fields inside symbol
        for item in node[1:]:
            if isinstance(item, list) and item:
                tag = item[0]
                if tag == Symbol('lib_id') and len(item) >= 2:
                    lib_id = item[1]
                elif tag == Symbol('in_bom') and len(item) >= 2:
                    in_bom = (item[1] == Symbol('yes'))
                elif tag == Symbol('exclude_from_sim') and len(item) >= 2:
                    exclude_from_sim = (item[1] == Symbol('yes'))
                elif tag == Symbol('dnp') and len(item) >= 2:
                    dnp = (item[1] == Symbol('yes'))
                elif tag == Symbol('property') and len(item) >= 3:
                    key = item[1]
                    val = item[2]
                    if key == 'Reference':
                        reference = val
                    elif key == 'Value':
                        value = val
                    elif key == 'Footprint':
                        footprint = val
                    else:
                        # capture interesting attributes
                        if isinstance(key, str) and isinstance(val, str):
                            properties[key] = val

        if not reference:
            return None
        return Component(
            reference=reference,
            lib_id=lib_id,
            value=value or "",
            footprint=footprint or "",
            properties=properties,
            in_bom=in_bom,
            exclude_from_sim=exclude_from_sim,
            dnp=dnp,
        )


class InventoryMatcher:
    """Matches components to inventory items"""
    
    def __init__(self, inventory_path: Path):
        self.inventory_path = inventory_path
        self.inventory: List[InventoryItem] = []
        self.inventory_fields: List[str] = []
        self._load_inventory()
    
    def _load_inventory(self):
        """Load inventory from supported file format (CSV, Excel, Numbers)"""
        file_extension = self.inventory_path.suffix.lower()
        
        if file_extension == '.csv':
            self._load_csv_inventory()
        elif file_extension in ['.xlsx', '.xls']:
            if not EXCEL_SUPPORT:
                raise ImportError("Excel support requires openpyxl package. Install with: pip install openpyxl")
            self._load_excel_inventory()
        elif file_extension == '.numbers':
            if not NUMBERS_SUPPORT:
                raise ImportError("Numbers support requires numbers-parser package. Install with: pip install numbers-parser")
            self._load_numbers_inventory()
        else:
            raise ValueError(f"Unsupported inventory file format: {file_extension}. Supported formats: .csv, .xlsx, .xls, .numbers")
    
    def _load_csv_inventory(self):
        """Load inventory from CSV file"""
        with open(self.inventory_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            headers = reader.fieldnames or []
            self._process_inventory_data(headers, rows)
    
    def _load_excel_inventory(self):
        """Load inventory from Excel file (.xlsx or .xls)"""
        workbook = openpyxl.load_workbook(self.inventory_path, data_only=True)
        # Use the first worksheet
        worksheet = workbook.active
        
        # Find the header row by looking for 'IPN' column
        header_row = None
        start_col = 1
        
        for row_num in range(1, min(10, worksheet.max_row + 1)):  # Check first 10 rows
            for col_num in range(1, min(10, worksheet.max_column + 1)):  # Check first 10 columns
                cell_value = worksheet.cell(row_num, col_num).value
                if cell_value and str(cell_value).strip().upper() == 'IPN':
                    header_row = row_num
                    start_col = col_num
                    break
            if header_row:
                break
        
        if not header_row:
            raise ValueError("Could not find 'IPN' header column in Excel file. Make sure the inventory has an 'IPN' column.")
        
        # Get headers from the identified header row
        headers = []
        col_num = start_col
        while col_num <= worksheet.max_column:
            cell_value = worksheet.cell(header_row, col_num).value
            if cell_value is not None and str(cell_value).strip():
                headers.append(str(cell_value).strip())
                col_num += 1
            else:
                # Stop when we hit an empty header cell
                break
        
        # Get data rows
        rows = []
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_offset, header in enumerate(headers):
                col_num = start_col + col_offset
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row_num, col_num).value
                    value_str = str(cell_value).strip() if cell_value is not None else ''
                    row_data[header] = value_str
                    if value_str:  # Check if this row has any data
                        has_data = True
                else:
                    row_data[header] = ''
            
            # Only add rows that have some data
            if has_data:
                rows.append(row_data)
        
        workbook.close()
        self._process_inventory_data(headers, rows)
    
    def _load_numbers_inventory(self):
        """Load inventory from Apple Numbers file"""
        doc = NumbersDocument(self.inventory_path)
        # Get the first table from the first sheet
        if not doc.sheets:
            raise ValueError("No sheets found in Numbers file")
        
        sheet = doc.sheets[0]
        if not sheet.tables:
            raise ValueError("No tables found in first sheet of Numbers file")
        
        table = sheet.tables[0]
        
        # Find the header row by looking for 'IPN' column (similar to Excel)
        header_row_idx = None
        start_col = None
        
        for row_idx in range(min(10, table.num_rows)):  # Check first 10 rows
            for col_idx in range(min(10, table.num_cols)):  # Check first 10 columns
                cell = table.cell(row_idx, col_idx)
                if cell.value and str(cell.value).strip().upper() == 'IPN':
                    header_row_idx = row_idx
                    start_col = col_idx
                    break
            if header_row_idx is not None:
                break
        
        if header_row_idx is None:
            raise ValueError("Could not find 'IPN' header column in Numbers file. Make sure the inventory has an 'IPN' column.")
        
        # Get headers from the identified header row
        headers = []
        
        for col_idx in range(start_col, table.num_cols):
            cell = table.cell(header_row_idx, col_idx)
            if cell.value is not None and str(cell.value).strip():
                headers.append(str(cell.value).strip())
            else:
                # Stop when we hit an empty header cell
                break
        
        # Get data rows
        rows = []
        for row_idx in range(header_row_idx + 1, table.num_rows):
            row_data = {}
            has_data = False
            
            for col_offset, header in enumerate(headers):
                col_idx = start_col + col_offset
                if col_idx < table.num_cols:
                    cell = table.cell(row_idx, col_idx)
                    value_str = str(cell.value).strip() if cell.value is not None else ''
                    row_data[header] = value_str
                    if value_str:  # Check if this row has any data
                        has_data = True
                else:
                    row_data[header] = ''
            
            # Only add rows that have some data
            if has_data:
                rows.append(row_data)
        
        self._process_inventory_data(headers, rows)
    
    def _process_inventory_data(self, headers: List[str], rows: List[Dict[str, str]]):
        """Process inventory data from any source format into InventoryItem objects"""
        # Validate required headers
        required_headers = ['IPN', 'Category']
        header_upper = [h.upper() for h in headers]
        missing_headers = [req for req in required_headers if req.upper() not in header_upper]
        
        if missing_headers:
            raise ValueError(
                f"Inventory file is missing required columns: {', '.join(missing_headers)}. "
                f"Found columns: {', '.join(headers)}"
            )
        
        # Clean up field names - replace newlines and extra whitespace
        self.inventory_fields = []
        for field in headers:
            if field:
                # Replace newlines with spaces and normalize whitespace
                clean_field = ' '.join(field.replace('\n', ' ').replace('\r', ' ').split())
                self.inventory_fields.append(clean_field)
        
        for row in rows:
            if not row.get('IPN'):  # Skip empty rows
                continue
            
            # No need to parse stocking info - Priority column handles all ranking
            item = InventoryItem(
                ipn=row.get('IPN', ''),
                keywords=row.get('Keywords', ''),
                category=row.get('Category', ''),
                description=row.get('Description', ''),
                smd=row.get('SMD', ''),
                value=row.get('Value', ''),
                type=row.get('Type', ''),
                tolerance=row.get('Tolerance', ''),
                voltage=row.get('V', ''),
                amperage=row.get('A', ''),
                wattage=row.get('W', ''),
                lcsc=row.get('LCSC', ''),
                manufacturer=row.get('Manufacturer', ''),
                mfgpn=row.get('MFGPN', ''),
                datasheet=row.get('Datasheet', ''),
                package=row.get('Package', ''),
                priority=self._parse_priority(row.get('Priority', str(DEFAULT_PRIORITY))),
                raw_data=row
            )
            self.inventory.append(item)
    
    def _parse_priority(self, priority_str: str) -> int:
        """Parse priority value from CSV, defaulting to DEFAULT_PRIORITY if invalid"""
        try:
            return int(priority_str.strip()) if priority_str.strip() else DEFAULT_PRIORITY
        except (ValueError, AttributeError):
            return DEFAULT_PRIORITY
    
    def find_matches(self, component: Component, debug: bool = False) -> List[Tuple[InventoryItem, int, Optional[str]]]:
        """Find matching inventory items for a component using primary filtering first."""
        matches: List[Tuple[InventoryItem, int, Tuple[int,int], Optional[str]]] = []
        
        # Primary filters: category/type, package, exact value
        comp_type = self._get_component_type(component)
        comp_pkg = self._extract_package_from_footprint(component.footprint)
        comp_val_norm = self._normalize_value(component.value) if component.value else ''
        
        debug_info = []
        if debug:
            debug_info.append(f"Component: {component.reference} ({component.lib_id})")
            debug_info.append(f"Detected type: {comp_type or 'Unknown'}")
            debug_info.append(f"Package: {comp_pkg or 'None'}")
            debug_info.append(f"Value: {component.value or 'None'}")
        
        candidates_checked = 0
        candidates_passed = 0
        
        for item in self.inventory:
            candidates_checked += 1
            if not self._passes_primary_filters(comp_type, comp_pkg, comp_val_norm, component, item, debug):
                continue
            candidates_passed += 1
            score, score_debug = self._calculate_match_score(component, item, debug)
            if score > 0:
                item_debug = None
                if debug:
                    item_debug = f"IPN: {item.ipn}, Score: {score}, Priority: {item.priority}" + (f", {score_debug}" if score_debug else "")
                matches.append((item, score, item.priority, item_debug))
        
        if debug:
            debug_info.append(f"Candidates: {candidates_checked}, Passed filters: {candidates_passed}, Matched: {len(matches)}")
        
        # Sort by priority first (lower is better), then by score desc
        matches.sort(key=lambda x: (x[2], -x[1]))
        
        # Format return with debug info
        result_debug = "; ".join(debug_info) if debug and debug_info else None
        return [(itm, sc, result_debug if i == 0 else item_debug) for i, (itm, sc, _, item_debug) in enumerate(matches)]
    
    def _calculate_match_score(self, component: Component, item: InventoryItem, debug: bool = False) -> Tuple[int, Optional[str]]:
        """Calculate match score between component and inventory item"""
        score = 0
        debug_parts = []
        
        # Component type matching
        comp_type = self._get_component_type(component)
        if comp_type and comp_type in item.category:
            score += 50
            if debug:
                debug_parts.append(f"Type match: +50 ({comp_type} in {item.category})")
        elif debug and comp_type:
            debug_parts.append(f"Type mismatch: {comp_type} not in {item.category}")
        
        # Value matching
        if component.value and self._values_match(component.value, item.value):
            score += 40
            if debug:
                debug_parts.append(f"Value match: +40 ({component.value} = {item.value})")
        elif debug and component.value:
            debug_parts.append(f"Value mismatch: {component.value} ≠ {item.value}")
        
        # Footprint matching  
        if component.footprint and item.package:
            footprint_match = self._footprint_matches(component.footprint, item.package)
            if footprint_match:
                score += 30
                if debug:
                    debug_parts.append(f"Footprint match: +30")
            elif debug:
                debug_parts.append(f"Footprint mismatch: {component.footprint} ≠ {item.package}")
        
        # Property matching
        prop_score = self._match_properties(component, item)
        if prop_score > 0:
            score += prop_score
            if debug:
                debug_parts.append(f"Property match: +{prop_score}")
        
        # Keyword matching
        if component.value in item.keywords:
            score += 10
            if debug:
                debug_parts.append(f"Keyword match: +10")
        
        debug_info = ", ".join(debug_parts) if debug and debug_parts else None
        return score, debug_info

    def _extract_package_from_footprint(self, footprint: str) -> str:
        fp = (footprint or '').lower()
        
        # Try direct matching with SMD packages (standard format)
        # Sort by length descending to match longer patterns first (e.g., 'sot-23' before 'sot')
        for pattern in sorted(PackageType.SMD_PACKAGES, key=len, reverse=True):
            if pattern in fp:
                return pattern
        
        return ''

    def _passes_primary_filters(self, comp_type: Optional[str], comp_pkg: str, comp_val_norm: str,
                                component: Component, item: InventoryItem, debug: bool = False) -> bool:
        # 1) Type/category must match if we could determine it
        if comp_type:
            cat = (item.category or '').upper()
            if comp_type not in cat:
                return False
        # 2) Package must match when we can extract it
        if comp_pkg:
            ipkg = (item.package or '').lower()
            if comp_pkg not in ipkg:
                return False
        # 3) Value match by type (numeric for RES/CAP/IND)
        if comp_val_norm:
            if comp_type == ComponentType.RESISTOR:
                comp_num = self._parse_res_to_ohms(component.value)
                inv_num = self._parse_res_to_ohms(item.value)
                if comp_num is None or inv_num is None or abs(comp_num - inv_num) > 1e-12:
                    return False
            elif comp_type == ComponentType.CAPACITOR:
                comp_num = self._parse_cap_to_farad(component.value)
                inv_num = self._parse_cap_to_farad(item.value)
                if comp_num is None or inv_num is None or abs(comp_num - inv_num) > 1e-18:
                    return False
            elif comp_type == ComponentType.INDUCTOR:
                comp_num = self._parse_ind_to_henry(component.value)
                inv_num = self._parse_ind_to_henry(item.value)
                if comp_num is None or inv_num is None or abs(comp_num - inv_num) > 1e-18:
                    return False
            else:
                inv_val_norm = self._normalize_value(item.value) if item.value else ''
                if not inv_val_norm or inv_val_norm != comp_val_norm:
                    return False
        return True

    
    def _get_component_type(self, component: Component) -> Optional[str]:
        """Determine component type from lib_id or footprint"""
        return get_component_type(component.lib_id, component.footprint)
    
    def _values_match(self, comp_value: str, inv_value: str) -> bool:
        """Check if component and inventory values match"""
        if not comp_value or not inv_value:
            return False
        
        # Normalize values
        comp_norm = self._normalize_value(comp_value)
        inv_norm = self._normalize_value(inv_value)
        
        return comp_norm == inv_norm
    
    def _normalize_value(self, value: str) -> str:
        """Normalize values for non-resistance comparisons (legacy fallback)."""
        value = (value or '').strip().lower()
        # Strip unit symbols and collapse whitespace
        value = re.sub(r'[Ωω]|ohm', '', value)
        value = value.replace('μ', 'u')
        value = re.sub(r'\s+', '', value)
        return value

    # ---- Resistance parsing / EIA formatting helpers ----
    _OHM_RE = re.compile(r"^\s*([0-9]*\.?[0-9]+)\s*([kKmMrR]?)\s*\+?\s*$")

    def _parse_res_to_ohms(self, s: str) -> Optional[float]:
        """Delegate to common.values.parse_res_to_ohms (kept for back-compat)."""
        from .common.values import parse_res_to_ohms
        return parse_res_to_ohms(s)

    def _ohms_to_eia(self, ohms: float, *, force_trailing_zero: bool = False) -> str:
        """Delegate to common.values.ohms_to_eia (kept for back-compat)."""
        from .common.values import ohms_to_eia
        return ohms_to_eia(ohms, force_trailing_zero=force_trailing_zero)

    # ---- Capacitor parsing / EIA-ish formatting ----
    def _parse_cap_to_farad(self, s: str) -> Optional[float]:
        from .common.values import parse_cap_to_farad
        return parse_cap_to_farad(s)

    def _cap_unit_multiplier(self, unit: str) -> float:
        from .common.values import cap_unit_multiplier
        return cap_unit_multiplier(unit)

    def _farad_to_eia(self, farad: float) -> str:
        from .common.values import farad_to_eia
        return farad_to_eia(farad)

    # ---- Inductor parsing / EIA-ish formatting ----
    def _parse_ind_to_henry(self, s: str) -> Optional[float]:
        from .common.values import parse_ind_to_henry
        return parse_ind_to_henry(s)

    def _ind_unit_multiplier(self, unit: str) -> float:
        from .common.values import ind_unit_multiplier
        return ind_unit_multiplier(unit)

    def _henry_to_eia(self, henry: float) -> str:
        from .common.values import henry_to_eia
        return henry_to_eia(henry)
    
    def _footprint_matches(self, footprint: str, package: str) -> bool:
        """Check if footprint matches package inventory designation"""
        if not footprint or not package:
            return False
        
        footprint = footprint.lower()
        package = package.lower()
        
        # First try direct matching: check if any SMD package pattern
        # appears in both footprint and package (most common case)
        for pattern in PackageType.SMD_PACKAGES:
            if pattern in footprint and pattern in package:
                return True
        
        # Second try: automatic dash removal for inventory naming variations
        # Many inventories use 'sot23' instead of 'sot-23', 'sod123' instead of 'sod-123', etc.
        for pattern in PackageType.SMD_PACKAGES:
            if '-' in pattern:
                pattern_no_dash = pattern.replace('-', '')
                if pattern in footprint and pattern_no_dash in package:
                    return True
        
        return False
    
    def _match_properties(self, component: Component, item: InventoryItem) -> int:
        """Match component properties with inventory item using category-specific logic"""
        score = 0
        
        # Get component type to determine which properties are relevant
        comp_type = self._get_component_type(component)
        relevant_fields = get_category_fields(comp_type) if comp_type else get_category_fields('')
        
        # Tolerance matching - exact match preferred; tighter tolerances substitute only when exact unavailable
        if CommonFields.TOLERANCE in relevant_fields and CommonFields.TOLERANCE in component.properties and item.tolerance:
            comp_tol = self._parse_tolerance_percent(component.properties[CommonFields.TOLERANCE])
            item_tol = self._parse_tolerance_percent(item.tolerance)
            
            if comp_tol is not None and item_tol is not None:
                if comp_tol == item_tol:
                    # Exact match - full score
                    score += ScoreWeights.TOLERANCE_EXACT
                elif item_tol < comp_tol:
                    # Inventory has tighter tolerance than required - acceptable substitution
                    # Prefer next-tighter over overly-tight: score inversely proportional to tightness gap
                    # Gap = comp_tol - item_tol (positive value)
                    # Closer gap (smaller tightness difference) gets higher score
                    # Use a bonus that decreases as tolerance gets tighter than necessary
                    tolerance_gap = comp_tol - item_tol
                    # Award full bonus if within 1% of required, half bonus if tighter
                    if tolerance_gap <= 1.0:
                        score += ScoreWeights.TOLERANCE_BETTER  # Nearly exact or 1% tighter
                    else:
                        score += max(1, ScoreWeights.TOLERANCE_BETTER // 2)  # Significantly tighter gets reduced bonus
                # If item_tol >= comp_tol, no points (exact match handled above, looser can't substitute)
        
        # Voltage matching (V or Voltage)
        if any(field in relevant_fields for field in [CommonFields.VOLTAGE, 'Voltage']) and item.voltage:
            voltage_props = [p for p in ['Voltage', CommonFields.VOLTAGE] if p in component.properties]
            for prop in voltage_props:
                if component.properties[prop] in item.voltage:
                    score += ScoreWeights.VOLTAGE_MATCH
                    break
        
        # Current/Amperage matching (A or Amperage)
        if any(field in relevant_fields for field in [CommonFields.AMPERAGE, 'Amperage']) and item.amperage:
            current_props = [p for p in [CommonFields.AMPERAGE, 'Amperage'] if p in component.properties]
            for prop in current_props:
                if component.properties[prop] in item.amperage:
                    score += ScoreWeights.CURRENT_MATCH
                    break
        
        # Power/Wattage matching (W, Power, or P)
        if any(field in relevant_fields for field in [CommonFields.WATTAGE, CommonFields.POWER]) and item.wattage:
            power_props = [p for p in [CommonFields.WATTAGE, CommonFields.POWER, 'P'] if p in component.properties]
            for prop in power_props:
                if component.properties[prop] in item.wattage:
                    score += ScoreWeights.POWER_MATCH
                    break
        
        # LED-specific properties
        if comp_type == ComponentType.LED:
            # Wavelength matching
            if 'Wavelength' in component.properties and hasattr(item, 'wavelength') and item.wavelength:
                if component.properties['Wavelength'] in item.wavelength:
                    score += ScoreWeights.LED_WAVELENGTH
            
            # Luminous intensity (mcd) matching
            if 'mcd' in component.properties and hasattr(item, 'mcd') and item.mcd:
                if component.properties['mcd'] in item.mcd:
                    score += ScoreWeights.LED_INTENSITY
            
            # Angle matching
            if 'Angle' in component.properties and hasattr(item, 'angle') and item.angle:
                if component.properties['Angle'] in item.angle:
                    score += ScoreWeights.LED_ANGLE
        
        # Oscillator-specific properties
        if comp_type == ComponentType.OSCILLATOR:
            if 'Frequency' in component.properties and hasattr(item, 'frequency') and item.frequency:
                if component.properties['Frequency'] in item.frequency:
                    score += ScoreWeights.OSC_FREQUENCY
            
            if 'Stability' in component.properties and hasattr(item, 'stability') and item.stability:
                if component.properties['Stability'] in item.stability:
                    score += ScoreWeights.OSC_STABILITY
            
            if 'Load' in component.properties and hasattr(item, 'load') and item.load:
                if component.properties['Load'] in item.load:
                    score += ScoreWeights.OSC_LOAD
        
        # Connector-specific properties
        if comp_type == ComponentType.CONNECTOR:
            if 'Pitch' in component.properties and hasattr(item, 'pitch') and item.pitch:
                if component.properties['Pitch'] in item.pitch:
                    score += ScoreWeights.CON_PITCH
        
        # MCU/IC-specific properties
        if comp_type in [ComponentType.MICROCONTROLLER, ComponentType.INTEGRATED_CIRCUIT]:
            if 'Family' in component.properties and hasattr(item, 'family') and item.family:
                if component.properties['Family'] in item.family:
                    score += ScoreWeights.MCU_FAMILY
        
        # Generic property matching for any additional properties
        for prop_name, prop_value in component.properties.items():
            if prop_name in relevant_fields and hasattr(item, prop_name.lower()) and getattr(item, prop_name.lower(), None):
                if prop_value in getattr(item, prop_name.lower()):
                    score += ScoreWeights.GENERIC_PROPERTY  # Lower score for generic matches
        
        return score
    
    def _parse_tolerance_percent(self, tol_str: str) -> Optional[float]:
        """Parse tolerance string like '±5%', '5%', '±1%' to numeric percentage"""
        if not tol_str:
            return None
        
        # Clean up the string - remove ±, %, spaces
        cleaned = tol_str.strip().replace('±', '').replace('%', '').strip()
        
        try:
            return float(cleaned)
        except ValueError:
            return None


class BOMGenerator:
    """Generates bill of materials from components and inventory matches"""
    
    def __init__(self, components: List[Component], matcher: InventoryMatcher):
        self.components = components
        self.matcher = matcher
    
    def generate_bom(self, verbose: bool = False, debug: bool = False, smd_only: bool = False) -> Tuple[List[BOMEntry], int, List[dict]]:
        """Generate bill of materials"""
        bom_entries: List[BOMEntry] = []
        debug_diagnostics: List[dict] = []
        
        # Group components by value and footprint
        grouped_components = self._group_components()
        
        for group_key, group_components in grouped_components.items():
            reference = group_components[0].reference
            quantity = len(group_components)
            
            # Find matches for first component in group
            matches = self.matcher.find_matches(group_components[0], debug=debug)
            
            if matches:
                # Use best match
                if debug:
                    best_item, score, match_debug = matches[0]
                else:
                    best_item, score, _ = matches[0]
                
                # Determine if schematic implies 1% (explicit trailing zero like 10k0 or Tolerance <=1%)
                comp0 = group_components[0]
                desired_1pct = False
                # explicit precision pattern (any trailing digit after unit indicates precision)
                if get_component_type(comp0.lib_id, comp0.footprint) == ComponentType.RESISTOR:
                    import re as _re
                    explicit_precision = bool(_re.match(r"^\s*\d+[kKmMrR]\d+\s*", comp0.value or ''))
                    tol_str = (comp0.properties.get(CommonFields.TOLERANCE) or '').strip().replace('%','')
                    tol_ok = False
                    try:
                        tol_ok = float(tol_str) <= PRECISION_THRESHOLD if tol_str else False
                    except ValueError:
                        tol_ok = False
                    desired_1pct = explicit_precision or tol_ok
                
                # Check inventory for any 1% option among candidates
                has_1pct_option = any(((itm.tolerance or '').strip().startswith('1%')) for itm,_,_ in matches)
                warn = ""
                if desired_1pct and not has_1pct_option:
                    best_tol = (best_item.tolerance or '').strip() or 'unknown'
                    warn = f" Warning: schematic implies 1% resistor but no 1% inventory item found (best tolerance {best_tol})."
                
                # Create BOM entry with tie handling
                display_value = self._format_display_value(comp0)
                base_notes, viable_alts = self._analyze_matches(matches, best_item, verbose)
                
                # Debug information is handled by verbose console output, not BOM notes
                notes_combined = (base_notes + warn).strip()
                entry = BOMEntry(
                    reference=", ".join([c.reference for c in group_components]),
                    quantity=quantity,
                    value=display_value,
                    footprint=comp0.footprint,
                    lcsc=best_item.lcsc,
                    manufacturer=best_item.manufacturer,
                    mfgpn=best_item.mfgpn,
                    description=best_item.description,
                    datasheet=best_item.datasheet,
                    smd=best_item.smd,
                    match_quality=f"Score: {score}",
                    notes=notes_combined,
                    priority=best_item.priority,
                )
                
                bom_entries.append(entry)
                
                # Add viable alternative matches only
                for additional_item, additional_score in viable_alts:
                    additional_entry = BOMEntry(
                        reference=f"ALT: {', '.join([c.reference for c in group_components])}",
                        quantity=quantity,
                        value=display_value,
                        footprint=group_components[0].footprint,
                        lcsc=additional_item.lcsc,
                        manufacturer=additional_item.manufacturer,
                        mfgpn=additional_item.mfgpn,
                        description=additional_item.description,
                        datasheet=additional_item.datasheet,
                        smd=additional_item.smd,
                        match_quality=f"Score: {additional_score}",
                        notes="Alternative match",
                        priority=additional_item.priority,
                    )
                    bom_entries.append(additional_entry)
            else:
                # No matches found - provide diagnostic information in debug mode
                comp0 = group_components[0]
                display_value = self._format_display_value(comp0)
                
                debug_notes = ""
                diagnostic_data = None
                if debug:
                    diagnostic_data = self._analyze_no_match_component(comp0)
                    debug_notes = self._format_diagnostic_for_bom(diagnostic_data)
                
                notes = "No inventory match found" + (debug_notes if debug_notes else "")
                
                entry = BOMEntry(
                    reference=", ".join([c.reference for c in group_components]),
                    quantity=quantity,
                    value=display_value,
                    footprint=comp0.footprint,
                    lcsc="",
                    manufacturer="",
                    mfgpn="",
                    description="",
                    datasheet="",
                    smd="",
                    match_quality="No match",
                    notes=notes
                )
                bom_entries.append(entry)
                
                # Only collect diagnostic for console output if component will be included in final BOM
                if debug and diagnostic_data:
                    # Create a temporary entry to check SMD status for filtering
                    temp_entry = entry
                    if not smd_only or self._is_smd_component(temp_entry):
                        debug_diagnostics.append(diagnostic_data)
        
        # Filter for SMD components only if requested
        excluded_count = 0
        if smd_only:
            original_count = len(bom_entries)
            # Filter BOM entries
            filtered_entries = []
            for entry in bom_entries:
                if self._is_smd_component(entry):
                    filtered_entries.append(entry)
            bom_entries = filtered_entries
            excluded_count = original_count - len(bom_entries)
        
        # Sort BOM entries by category and component numbering
        bom_entries.sort(key=self._bom_sort_key)
        return bom_entries, excluded_count, debug_diagnostics
    
    def _is_smd_component(self, entry: BOMEntry) -> bool:
        """Check if a BOM entry represents an SMD component based on inventory data"""
        # Check the SMD field from the matched inventory item
        smd_field = (entry.smd or '').strip().upper()
        
        # Explicit SMD marking
        if smd_field in SMDType.SMD_VALUES:
            return True
        
        # Explicit non-SMD marking
        elif smd_field in SMDType.PTH_VALUES:
            return False
        
        # For unclear/empty SMD field, try to infer from footprint
        elif not smd_field or smd_field in SMDType.UNKNOWN_VALUES:
            footprint = (entry.footprint or '').lower()
            
            # Check for SMD package indicators in footprints
            if any(indicator in footprint for indicator in PackageType.SMD_PACKAGES):
                return True
            # Check for through-hole indicators
            elif any(indicator in footprint for indicator in PackageType.THROUGH_HOLE_PACKAGES):
                return False
            
            # For SMD filtering: if uncertain, exclude (strict SMD-only)
            return False
        
        else:
            # Unknown/unexpected SMD field values (like "Q16", "R12" etc.)
            # These are likely data errors or non-SMD related fields
            import sys
            print(f"Warning: Unexpected SMD field value '{smd_field}' for component {entry.reference} - treating as non-SMD", 
                  file=sys.stderr)
            return False
    
    def _analyze_no_match_component(self, component: Component) -> dict:
        """Analyze a component with no inventory matches and return structured diagnostic data"""
        # Component analysis
        comp_type = get_component_type(component.lib_id, component.footprint)
        comp_pkg = self.matcher._extract_package_from_footprint(component.footprint)
        comp_val_norm = self.matcher._normalize_value(component.value) if component.value else ''
        
        # Check for candidates by value and type without package filtering
        value_matches = 0
        type_matches = 0
        package_mismatches = []
        
        for item in self.matcher.inventory:
            # Check type matching
            if comp_type and comp_type in (item.category or '').upper():
                type_matches += 1
                
                # Check value matching for same type
                if comp_val_norm and self.matcher._values_match(component.value, item.value):
                    value_matches += 1
                    
                    # Check if package is the issue
                    if comp_pkg:
                        item_pkg = (item.package or '').lower()
                        if comp_pkg not in item_pkg:
                            package_mismatches.append((item, comp_pkg, item.package))
        
        # Determine issue type and details
        if comp_type:
            if type_matches == 0:
                issue_type = DiagnosticIssue.NO_TYPE_MATCH
                issue_details = {'comp_type': comp_type}
            elif value_matches == 0 and component.value:
                issue_type = DiagnosticIssue.NO_VALUE_MATCH
                issue_details = {'comp_type': comp_type, 'value': component.value}
            elif package_mismatches and comp_pkg:
                available_packages = set(item.package for item, _, _ in package_mismatches if item.package)
                if available_packages:
                    issue_type = DiagnosticIssue.PACKAGE_MISMATCH
                    issue_details = {
                        'value': component.value,
                        'available_packages': sorted(available_packages),
                        'required_package': comp_pkg
                    }
                else:
                    issue_type = DiagnosticIssue.PACKAGE_MISMATCH_GENERIC
                    issue_details = {'required_package': comp_pkg}
            else:
                issue_type = DiagnosticIssue.NO_MATCH
                issue_details = {}
        else:
            issue_type = DiagnosticIssue.TYPE_UNKNOWN
            issue_details = {}
        
        return {
            'component': {
                'reference': component.reference,
                'lib_id': component.lib_id,
                'value': component.value,
                'footprint': component.footprint
            },
            'analysis': {
                'type': comp_type,
                'package': comp_pkg,
                'value_normalized': comp_val_norm
            },
            'issue': {
                'type': issue_type,
                'details': issue_details
            }
        }
    
    def _generate_diagnostic_message(self, diagnostic_data: dict, format_type: str) -> str:
        """Generate diagnostic message from structured data for different output formats.
        
        Both formats contain the same diagnostic information, just formatted differently:
        - BOM format: semicolon-separated with DEBUG prefix for CSV compatibility
        - Console format: user-friendly multi-line format for readability
        
        Args:
            diagnostic_data: Structured diagnostic data
            format_type: 'bom' for BOM file format, 'console' for user-friendly console format
        """
        comp = diagnostic_data['component']
        analysis = diagnostic_data['analysis']
        issue = diagnostic_data['issue']
        
        if format_type == 'bom':
            # BOM format: concise semicolon-separated format for CSV compatibility
            lib_namespace = ''
            if ':' in comp['lib_id']:
                lib_namespace, _ = comp['lib_id'].split(':', 1)
            
            # Use concise component description like console format
            if not analysis['type']:
                lib_part = comp['lib_id'].split(':', 1)[1] if ':' in comp['lib_id'] else comp['lib_id']
                comp_desc = f"Component: {comp['reference']} ({comp['lib_id']}) from {lib_namespace} (part: {lib_part})"
            else:
                type_names = {
                    ComponentType.RESISTOR: 'Resistor', ComponentType.CAPACITOR: 'Capacitor', 
                    ComponentType.INDUCTOR: 'Inductor', ComponentType.DIODE: 'Diode', 
                    ComponentType.LED: 'LED', ComponentType.INTEGRATED_CIRCUIT: 'IC', 
                    ComponentType.CONNECTOR: 'Connector', ComponentType.SWITCH: 'Switch', 
                    ComponentType.TRANSISTOR: 'Transistor'
                }
                type_name = type_names.get(analysis['type'], analysis['type'])
                package_text = f" {analysis['package']}" if analysis['package'] else ''
                value_text = f" {comp['value']}" if comp['value'] else ''
                comp_desc = f"Component: {comp['reference']} ({comp['lib_id']}) is a{value_text}{package_text} {type_name}"
            
            # Generate issue message
            issue_msg = self._format_issue_message(issue, analysis.get('type'))
            
            return f" {comp_desc}; Issue: {issue_msg}"
        
        elif format_type == 'console':
            # Console format: concise user-friendly format with same core information
            lib_namespace = ''
            if ':' in comp['lib_id']:
                lib_namespace, _ = comp['lib_id'].split(':', 1)
            
            # Format main component description (contains same info as BOM format)
            if not analysis['type']:
                lib_part = comp['lib_id'].split(':', 1)[1] if ':' in comp['lib_id'] else comp['lib_id']
                main_desc = f"Component {comp['reference']} from {lib_namespace} (part: {lib_part})"
            else:
                type_names = {
                    ComponentType.RESISTOR: 'Resistor', ComponentType.CAPACITOR: 'Capacitor', 
                    ComponentType.INDUCTOR: 'Inductor', ComponentType.DIODE: 'Diode', 
                    ComponentType.LED: 'LED', ComponentType.INTEGRATED_CIRCUIT: 'IC', 
                    ComponentType.CONNECTOR: 'Connector', ComponentType.SWITCH: 'Switch', 
                    ComponentType.TRANSISTOR: 'Transistor'
                }
                type_name = type_names.get(analysis['type'], analysis['type'])
                package_text = f" {analysis['package']}" if analysis['package'] else ''
                value_text = f" {comp['value']}" if comp['value'] else ''
                main_desc = f"Component {comp['reference']} from {lib_namespace} is a{value_text}{package_text} {type_name}"
            
            # Generate issue message
            issue_msg = self._format_issue_message(issue, analysis.get('type'), format_type='console')
            
            return f"{main_desc}\n    Issue: {issue_msg}"
        
        else:
            raise ValueError(f"Unknown format_type: {format_type}")
    
    def _format_issue_message(self, issue: dict, comp_type: str, format_type: str = 'bom') -> str:
        """Format the issue message based on issue type and output format."""
        issue_type = issue['type']
        details = issue['details']
        
        if issue_type == DiagnosticIssue.TYPE_UNKNOWN:
            if format_type == 'console':
                return "Cannot determine component type - may be a non-electronic part (board outline, label, etc.)"
            else:
                return "Component type could not be determined"
        
        elif issue_type == DiagnosticIssue.NO_TYPE_MATCH:
            comp_type_name = details['comp_type']
            if format_type == 'console':
                type_names = {
                    ComponentType.RESISTOR: 'resistor', ComponentType.CAPACITOR: 'capacitor', 
                    ComponentType.INDUCTOR: 'inductor', ComponentType.DIODE: 'diode', 
                    ComponentType.LED: 'led', ComponentType.INTEGRATED_CIRCUIT: 'ic', 
                    ComponentType.CONNECTOR: 'connector', ComponentType.SWITCH: 'switch', 
                    ComponentType.TRANSISTOR: 'transistor'
                }
                friendly_name = type_names.get(comp_type_name, comp_type_name.lower())
                return f"No {friendly_name}s in inventory"
            else:
                return f"No {comp_type_name} components found in inventory"
        
        elif issue_type == DiagnosticIssue.NO_VALUE_MATCH:
            comp_type_name = details['comp_type']
            value = details['value']
            if format_type == 'console':
                type_names = {
                    ComponentType.RESISTOR: 'resistor', ComponentType.CAPACITOR: 'capacitor', 
                    ComponentType.INDUCTOR: 'inductor', ComponentType.DIODE: 'diode', 
                    ComponentType.LED: 'led', ComponentType.INTEGRATED_CIRCUIT: 'ic', 
                    ComponentType.CONNECTOR: 'connector', ComponentType.SWITCH: 'switch', 
                    ComponentType.TRANSISTOR: 'transistor'
                }
                friendly_name = type_names.get(comp_type_name, comp_type_name.lower())
                return f"No {friendly_name}s with value '{value}' in inventory"
            else:
                return f"No {comp_type_name} components with value {value} found"
        
        elif issue_type == DiagnosticIssue.PACKAGE_MISMATCH:
            available = ', '.join(details['available_packages'])
            required = details['required_package']
            value = details['value']
            return f"Value '{value}' available in {available} packages, but not {required}"
        
        elif issue_type == DiagnosticIssue.PACKAGE_MISMATCH_GENERIC:
            required = details['required_package']
            if format_type == 'console':
                return f"Package mismatch - needs {required}"
            else:
                return f"Package mismatch - required {required}"
        
        else:  # no_match
            return "Component specification doesn't match any inventory items"
    
    def _format_diagnostic_for_bom(self, diagnostic_data: dict) -> str:
        """Format diagnostic data for BOM file output (with DEBUG prefix)"""
        return self._generate_diagnostic_message(diagnostic_data, 'bom')
    
    def _generate_no_match_diagnostics(self, component: Component) -> str:
        """Generate diagnostic information for components with no inventory matches"""
        diagnostic_data = self._analyze_no_match_component(component)
        return self._format_diagnostic_for_bom(diagnostic_data)
    
    def _analyze_matches(self, matches: List[Tuple[InventoryItem, int, Optional[str]]], best_item: InventoryItem, verbose: bool) -> Tuple[str, List[Tuple[InventoryItem, int]]]:
        """Handle ties: arbitrary choice by default, show ties only with verbose flag"""
        if len(matches) <= 1:
            return "", []
        
        best_priority = best_item.priority
        tied_items = []
        
        # Find items that tie with the best priority
        for item, score, _ in matches[1:]:  # Skip the best match, ignore debug info
            if item.priority == best_priority:
                tied_items.append((item, score))
        
        # Handle ties based on verbose flag
        if tied_items:
            if verbose:
                # Show ties in verbose mode for debugging/transparency
                total_tied = len(tied_items) + 1  # +1 for the best match
                notes = f"Tied priority {best_priority}: {total_tied} options"
                # Limit ALT entries to keep BOM manageable
                return notes, tied_items[:2]
            else:
                # Default: arbitrary choice (use first match), no ALT entries
                return "", []
        else:
            # No ties - single best choice
            return "", []
    
    def _bom_sort_key(self, entry: BOMEntry) -> Tuple[str, int, str]:
        """Generate sort key for BOM entry: (category, min_component_number, full_reference)"""
        refs = entry.reference.replace('ALT: ', '').split(', ')
        
        # Extract category and numbers from references
        categories = set()
        min_number = float('inf')
        
        for ref in refs:
            ref = ref.strip()
            # Extract category (letter prefix) and number
            category, number = self._parse_reference(ref)
            if category:
                categories.add(category)
            if number < min_number:
                min_number = number
        
        # Use primary category (first alphabetically if mixed)
        primary_category = sorted(categories)[0] if categories else 'Z'
        
        # Handle special case where min_number is still inf (no numbers found)
        if min_number == float('inf'):
            min_number = 0
        
        return (primary_category, int(min_number), entry.reference)
    
    def _parse_reference(self, ref: str) -> Tuple[str, float]:
        """Parse reference into category and number: R10 -> ('R', 10), LED4 -> ('LED', 4)"""
        if not ref:
            return '', float('inf')
        
        # Handle multi-letter prefixes (LED, etc.) and single letters (R, C, etc.)
        import re
        match = re.match(r'^([A-Za-z]+)(\d+)$', ref.strip())
        if match:
            category = match.group(1).upper()
            number = float(match.group(2))
            return category, number
        
        # Fallback for non-standard references
        return ref[0].upper() if ref else '', float('inf')
    
    def _format_display_value(self, component: Component) -> str:
        # Use EIA-like for R/C/L when possible
        comp_type = get_component_type(component.lib_id, component.footprint)
        if comp_type == ComponentType.RESISTOR:
            ohms = self.matcher._parse_res_to_ohms(component.value)
            if ohms is not None:
                tol = (component.properties.get(CommonFields.TOLERANCE) or '').strip().replace('%','')
                force0 = False
                try:
                    force0 = float(tol) <= PRECISION_THRESHOLD if tol else False
                except ValueError:
                    force0 = False
                # If schematic explicitly used trailing digit (e.g., 10K0, 47K5), preserve precision intent
                explicit_precision = bool(re.match(r"^\s*\d+[kKmMrR]\d+\s*", component.value or ''))
                return self.matcher._ohms_to_eia(ohms, force_trailing_zero=(force0 or explicit_precision))
        if comp_type == ComponentType.CAPACITOR:
            f = self.matcher._parse_cap_to_farad(component.value)
            if f is not None:
                return self.matcher._farad_to_eia(f)
        if comp_type == ComponentType.INDUCTOR:
            h = self.matcher._parse_ind_to_henry(component.value)
            if h is not None:
                return self.matcher._henry_to_eia(h)
        return component.value or ''

    def _group_components(self) -> Dict[str, List[Component]]:
        """Group components by their best matching inventory item"""
        groups = {}
        
        for component in self.components:
            # Find the best matching inventory item for this component
            matches = self.matcher.find_matches(component)
            
            if matches:
                # Use the IPN (Internal Part Number) of the best match as the group key
                best_item = matches[0][0]
                key = f"{best_item.ipn}_{component.footprint}"
            else:
                # No matches found - group by original value and footprint as fallback
                key = f"NO_MATCH_{component.value}_{component.footprint}"
            
            if key not in groups:
                groups[key] = []
            
            groups[key].append(component)
        
        return groups
    
    def get_available_fields(self, components: List[Component]) -> Dict[str, str]:
        """Get all available fields from BOM entries, inventory, and components with descriptions.
        
        All field names are normalized to snake_case internally.
        Returns dict mapping normalized field names to descriptions.
        """
        fields = {}
        
        # Standard BOM entry fields (normalized to snake_case)
        bom_fields = {
            'reference': 'Component reference designators (R1, C2, etc.)',
            'quantity': 'Number of components',
            'description': 'Component description from inventory',
            'value': 'Component value (10k, 100nF, etc.)',
            'footprint': 'PCB footprint name',
            'lcsc': 'LCSC part number',
            'manufacturer': 'Component manufacturer',
            'mfgpn': 'Manufacturer part number',
            'datasheet': 'Link to component datasheet',
            'smd': 'Surface mount/through-hole indicator',
            'match_quality': 'Inventory matching score (verbose mode)',
            'notes': 'Additional notes and warnings',
            'priority': 'Inventory item priority (verbose mode)'
        }
        fields.update(bom_fields)
        
        # Gather component properties from actual components - normalize field names
        component_props = set()
        for component in components:
            for prop_name in component.properties.keys():
                normalized = normalize_field_name(prop_name)
                component_props.add(normalized)
        
        # Normalize inventory field names
        inventory_names = set()
        for inv_field in self.matcher.inventory_fields:
            normalized = normalize_field_name(inv_field)
            inventory_names.add(normalized)
        
        # Create sets for systematic handling
        standard_field_names = set(fields.keys())  # Already normalized
        
        # Process all inventory and component fields systematically
        all_field_names = inventory_names.union(component_props)
        
        for field_name in sorted(all_field_names):
            # Skip if it's already a standard BOM field
            if field_name in standard_field_names:
                continue
                
            has_inventory = field_name in inventory_names
            has_component = field_name in component_props
            
            if has_inventory and has_component:
                # Ambiguous field - add unprefixed version and prefixed versions
                fields[field_name] = f'Ambiguous field: {field_to_header(field_name)} (will show both inventory and component versions)'
                fields[f'i:{field_name}'] = f'Inventory field: {field_to_header(field_name)}'
                fields[f'c:{field_name}'] = f'Component property: {field_to_header(field_name)}'
            elif has_inventory:
                # Inventory only - add both unprefixed and prefixed
                fields[field_name] = f'Inventory field: {field_to_header(field_name)}'
                fields[f'i:{field_name}'] = f'Inventory field: {field_to_header(field_name)}'
            elif has_component:
                # Component only - add both unprefixed and prefixed
                fields[field_name] = f'Component property: {field_to_header(field_name)}'
                fields[f'c:{field_name}'] = f'Component property: {field_to_header(field_name)}'
        
        return fields
    
    def _get_inventory_field_value(self, field: str, inventory_item: Optional[InventoryItem]) -> str:
        """Get a value from inventory item's raw data, handling cleaned field names.
        
        Field should be in normalized snake_case format.
        """
        if not inventory_item:
            return ''
        
        # Try to find a raw field that matches when normalized
        for raw_field, value in inventory_item.raw_data.items():
            if raw_field:
                # Clean up the raw field name (handle newlines) and normalize it
                cleaned_field = ' '.join(raw_field.replace('\n', ' ').replace('\r', ' ').split())
                if normalize_field_name(cleaned_field) == field:
                    return value
        
        return ''
    
    def _has_inventory_field(self, field: str, inventory_item: Optional[InventoryItem]) -> bool:
        """Check if field exists in inventory data.
        
        Field should be in normalized snake_case format.
        """
        if not inventory_item:
            return False
        
        # Check if a raw field matches when normalized
        for raw_field in inventory_item.raw_data.keys():
            if raw_field:
                # Clean up the raw field name (handle newlines) and normalize it
                cleaned_field = ' '.join(raw_field.replace('\n', ' ').replace('\r', ' ').split())
                if normalize_field_name(cleaned_field) == field:
                    return True
        
        return False
    
    def _get_field_value(self, field: str, entry: BOMEntry, component: Component, inventory_item: Optional[InventoryItem]) -> str:
        """Get the value for a specific field from BOM entry, component, or inventory.
        
        Expects field to be in normalized snake_case format (e.g., 'match_quality', 'i:package').
        """
        # Ensure field is normalized
        field = normalize_field_name(field)
        
        # Standard BOM entry fields (all normalized snake_case)
        if field == 'reference':
            return entry.reference
        elif field == 'quantity':
            return str(entry.quantity)
        elif field == 'description':
            return entry.description
        elif field == 'value':
            return entry.value
        elif field == 'footprint':
            return entry.footprint
        elif field == 'lcsc':
            return entry.lcsc
        elif field == 'manufacturer':
            return entry.manufacturer
        elif field == 'mfgpn':
            return entry.mfgpn
        elif field == 'datasheet':
            return entry.datasheet
        elif field == 'smd':
            return entry.smd
        elif field == 'match_quality':
            return entry.match_quality
        elif field == 'notes':
            return entry.notes
        elif field == 'priority':
            return str(entry.priority)
        
        # Component properties (prefixed with c:)
        elif field.startswith('c:'):
            prop_name = field[2:]  # Remove 'c:' prefix
            # Find matching property in component (normalized)
            for comp_prop, value in component.properties.items():
                if normalize_field_name(comp_prop) == prop_name:
                    return value
            return ''
        
        # Inventory fields (prefixed with i:)
        elif field.startswith('i:'):
            inv_field = field[2:]  # Remove 'i:' prefix
            return self._get_inventory_field_value(inv_field, inventory_item)
        
        # Ambiguous fields (no prefix) - check if it exists in both inventory and component
        else:
            if inventory_item:
                # Check if this field exists in both sources (using normalized names)
                has_inventory = self._has_inventory_field(field, inventory_item)
                has_component = any(normalize_field_name(prop) == field for prop in component.properties.keys())
                
                if has_inventory and has_component:
                    # Return both values with headers - this will be handled specially in CSV writing
                    inv_val = self._get_inventory_field_value(field, inventory_item)
                    # Find component value with matching normalized name
                    comp_val = ''
                    for prop_name, prop_val in component.properties.items():
                        if normalize_field_name(prop_name) == field:
                            comp_val = prop_val
                            break
                    return f"i:{inv_val}|c:{comp_val}"
                elif has_inventory:
                    return self._get_inventory_field_value(field, inventory_item)
                elif has_component:
                    # Find component value with matching normalized name
                    for prop_name, prop_val in component.properties.items():
                        if normalize_field_name(prop_name) == field:
                            return prop_val
                    return ''
            
            # Fallback: try inventory field
            if inventory_item:
                return self._get_inventory_field_value(field, inventory_item)
        
        return ''
    
    def write_bom_csv(self, bom_entries: List[BOMEntry], output_path: Path, fields: List[str]):
        """Write BOM entries to CSV file or stdout using the specified field list.
        
        Fields are expected to be in normalized snake_case format.
        CSV headers will be converted to human-readable Title Case format.
        
        Special output_path values for stdout:
        - "-"
        - "console"
        - "stdout"
        """
        import sys
        
        # Check if output should go to stdout
        output_str = str(output_path)
        use_stdout = output_str in ('-', 'console', 'stdout')
        
        if use_stdout:
            f = sys.stdout
        else:
            f = open(output_path, 'w', newline='', encoding='utf-8')
        
        try:
            writer = csv.writer(f)
            
            # Process fields to handle ambiguous ones
            header = []
            normalized_fields = []  # Keep normalized versions for value retrieval
            for field in fields:
                # Check if this is an ambiguous field by testing with a sample entry
                if bom_entries:
                    sample_entry = bom_entries[0]
                    first_ref = sample_entry.reference.replace('ALT: ', '').split(', ')[0]
                    sample_component = None
                    sample_inventory = None
                    
                    # Find sample component and inventory item
                    for comp in self.components:
                        if comp.reference == first_ref:
                            sample_component = comp
                            break
                    if sample_entry.lcsc:
                        for item in self.matcher.inventory:
                            if item.lcsc == sample_entry.lcsc:
                                sample_inventory = item
                                break
                    
                    # Test if field returns ambiguous value
                    if sample_component and sample_inventory:
                        test_value = self._get_field_value(field, sample_entry, sample_component, sample_inventory)
                        if '|' in test_value and test_value.startswith('i:') and 'c:' in test_value:
                            # This is an ambiguous field - split into two columns
                            header.extend([field_to_header(f'i:{field}'), field_to_header(f'c:{field}')])
                            normalized_fields.extend([f'i:{field}', f'c:{field}'])
                            continue
                
                # Regular field - convert to Title Case header
                header.append(field_to_header(field))
                normalized_fields.append(field)
            
            writer.writerow(header)
            
            # Write entries
            for entry in bom_entries:
                # Parse first component reference to get original component data
                first_ref = entry.reference.replace('ALT: ', '').split(', ')[0]
                component = None
                inventory_item = None
                
                # Find the component by reference
                for comp in self.components:
                    if comp.reference == first_ref:
                        component = comp
                        break
                
                # Find matching inventory item if LCSC is available
                if entry.lcsc:
                    for item in self.matcher.inventory:
                        if item.lcsc == entry.lcsc:
                            inventory_item = item
                            break
                
                # Build row using specified fields (normalized_fields)
                row = []
                i = 0
                while i < len(normalized_fields):
                    field = normalized_fields[i]
                    
                    # Check if this is a split ambiguous field pair
                    if (field.startswith('i:') and i + 1 < len(normalized_fields) and 
                        normalized_fields[i + 1].startswith('c:') and 
                        field[2:] == normalized_fields[i + 1][2:]):
                        
                        # Handle split ambiguous field
                        base_field = field[2:]  # Remove i: prefix
                        inv_value = self._get_inventory_field_value(base_field, inventory_item)
                        comp_value = ''
                        if component:
                            # Find component value with matching normalized name
                            for prop_name, prop_val in component.properties.items():
                                if normalize_field_name(prop_name) == base_field:
                                    comp_value = prop_val
                                    break
                        row.extend([inv_value, comp_value])
                        i += 2  # Skip the next field since we handled both
                    else:
                        # Regular field
                        value = self._get_field_value(field, entry, component or Component('', '', '', ''), inventory_item)
                        # Handle ambiguous values that weren't split in header
                        if '|' in value and value.startswith('i:') and 'c:' in value:
                            # Split the combined value
                            parts = value.split('|')
                            inv_part = parts[0][2:] if parts[0].startswith('i:') else ''
                            comp_part = parts[1][2:] if len(parts) > 1 and parts[1].startswith('c:') else ''
                            row.append(f"{inv_part} / {comp_part}")
                        else:
                            row.append(value)
                        i += 1
                
                writer.writerow(row)
        finally:
            if not use_stdout:
                f.close()


# Import file discovery functions from common.utils
from .common.utils import (
    find_best_schematic,
    is_hierarchical_schematic,
    extract_sheet_files,
    process_hierarchical_schematic,
)


def print_debug_diagnostics(diagnostics: List[dict]):
    """Print debug diagnostics in a concise, user-friendly format."""
    if not diagnostics:
        return
    
    print()
    print("Warnings:")
    print("=" * 60)
    
    for i, diagnostic_data in enumerate(diagnostics, 1):
        formatted_message = _format_diagnostic_for_console(diagnostic_data)
        print(f"{i:2d}. {formatted_message}")
        if i < len(diagnostics):  # Add blank line between diagnostics except after the last one
            print()
    
    print()


def _format_diagnostic_for_console(diagnostic_data: dict) -> str:
    """Format structured diagnostic data for console output."""
    # Use the same instance method that handles console formatting
    # Need to create a temporary instance since this is a standalone function
    # Create a minimal BOMGenerator instance for method access
    temp_generator = BOMGenerator([], None)
    return temp_generator._generate_diagnostic_message(diagnostic_data, 'console')


def _shorten_url(url: str, max_length: int = 30) -> str:
    """Shorten URLs for better table display."""
    if not url or len(url) <= max_length:
        return url
    
    # For HTTPS URLs, show domain + path start + end
    if url.startswith('https://'):
        # Extract domain and path
        parts = url[8:].split('/', 1)  # Remove https://
        domain = parts[0]
        path = '/' + parts[1] if len(parts) > 1 else ''
        
        if len(domain) > max_length - 6:  # 6 chars for "..." + some path
            return domain[:max_length-3] + '...'
        
        if len(domain + path) <= max_length:
            return domain + path
        
        # Show domain + start/end of path
        remaining = max_length - len(domain) - 6  # 6 for "/.../"
        if remaining > 0:
            path_start = path[1:remaining//2] if path.startswith('/') else path[:remaining//2]
            path_end = path[-(remaining//2):] if remaining//2 > 0 else ''
            return f"{domain}/.../{path_end}" if path_end else f"{domain}/..."
        else:
            return domain + '/...'
    
    # For other URLs, just truncate with ellipsis
    return url[:max_length-3] + '...' if len(url) > max_length else url


def _wrap_text(text: str, width: int) -> List[str]:
    """Wrap text to fit within specified width, breaking at word boundaries when possible."""
    if not text or width <= 0:
        return ['']
    
    if len(text) <= width:
        return [text]
    
    lines = []
    words = text.split()
    current_line = ''
    
    for word in words:
        # Truncate word if it's too long to fit in any line
        if len(word) > width:
            word = word[:width-3] + '...'
        
        # If adding this word would exceed width
        if current_line and len(current_line + ' ' + word) > width:
            # If current line has content, save it and start new line
            lines.append(current_line)
            current_line = word
        else:
            # Add word to current line
            if current_line:
                current_line += ' ' + word
            else:
                current_line = word
    
    # Add any remaining content
    if current_line:
        lines.append(current_line)
    
    return lines if lines else ['']


def print_bom_table(bom_entries: List[BOMEntry], verbose: bool = False, include_mfg: bool = False):
    """Print BOM entries as a formatted console table with word wrapping and URL shortening."""
    if not bom_entries:
        print("No BOM entries to display.")
        return
    
    # Determine columns to display
    headers = ['Reference', 'Qty', 'Value', 'Footprint', 'LCSC']
    if include_mfg:
        headers.extend(['Manufacturer', 'MFGPN'])
    headers.extend(['Datasheet', 'SMD'])
    if verbose:
        headers.extend(['Match_Quality', 'Priority'])
    
    # Check if any entries have notes
    any_notes = any((e.notes or '').strip() for e in bom_entries)
    if any_notes:
        headers.append('Notes')
    
    # Set maximum column widths for better table layout
    max_widths = {
        'Reference': 60,  # Allow long reference lists
        'Qty': 5,
        'Value': 12,
        'Footprint': 20,
        'LCSC': 10,
        'Manufacturer': 15,
        'MFGPN': 18,
        'Datasheet': 35,  # URLs get special handling
        'SMD': 4,
        'Match_Quality': 15,
        'Priority': 8,
        'Notes': 50  # Allow reasonable space for notes
    }
    
    # Calculate optimal column widths
    col_widths = {}
    for header in headers:
        col_widths[header] = len(header)
    
    # Calculate widths based on data, respecting maximums
    for entry in bom_entries:
        col_widths['Reference'] = min(max_widths['Reference'], max(col_widths['Reference'], len(entry.reference)))
        col_widths['Qty'] = min(max_widths['Qty'], max(col_widths['Qty'], len(str(entry.quantity))))
        col_widths['Value'] = min(max_widths['Value'], max(col_widths['Value'], len(entry.value)))
        col_widths['Footprint'] = min(max_widths['Footprint'], max(col_widths['Footprint'], len(entry.footprint)))
        col_widths['LCSC'] = min(max_widths['LCSC'], max(col_widths['LCSC'], len(entry.lcsc)))
        if include_mfg:
            col_widths['Manufacturer'] = min(max_widths['Manufacturer'], max(col_widths['Manufacturer'], len(entry.manufacturer)))
            col_widths['MFGPN'] = min(max_widths['MFGPN'], max(col_widths['MFGPN'], len(entry.mfgpn)))
        # Datasheet width based on shortened URLs
        shortened_url = _shorten_url(entry.datasheet, max_widths['Datasheet'])
        col_widths['Datasheet'] = min(max_widths['Datasheet'], max(col_widths['Datasheet'], len(shortened_url)))
        col_widths['SMD'] = min(max_widths['SMD'], max(col_widths['SMD'], len(entry.smd)))
        if verbose:
            col_widths['Match_Quality'] = min(max_widths['Match_Quality'], max(col_widths['Match_Quality'], len(entry.match_quality)))
            col_widths['Priority'] = min(max_widths['Priority'], max(col_widths['Priority'], len(str(entry.priority))))
        if any_notes:
            # Notes width based on first line of wrapped text
            notes_lines = _wrap_text(entry.notes or '', max_widths['Notes'])
            first_line_len = len(notes_lines[0]) if notes_lines else 0
            col_widths['Notes'] = min(max_widths['Notes'], max(col_widths['Notes'], first_line_len))
    
    # Print header
    header_line = ''
    separator_line = ''
    for i, header in enumerate(headers):
        width = col_widths[header]
        header_line += f"{header:<{width}}"
        separator_line += '-' * width
        if i < len(headers) - 1:
            header_line += ' | '
            separator_line += '-+-'
    
    print()
    print("BOM Table:")
    print("=" * min(120, len(header_line)))
    print(header_line)
    print(separator_line)
    
    # Print entries with word wrapping support
    for entry in bom_entries:
        # Prepare all cell content with wrapping
        cell_lines = {}
        max_lines = 1
        
        for header in headers:
            width = col_widths[header]
            
            if header == 'Reference':
                lines = _wrap_text(entry.reference, width)
            elif header == 'Qty':
                lines = [str(entry.quantity)]
            elif header == 'Value':
                lines = _wrap_text(entry.value, width)
            elif header == 'Footprint':
                lines = _wrap_text(entry.footprint, width)
            elif header == 'LCSC':
                lines = [entry.lcsc]
            elif header == 'Manufacturer':
                lines = _wrap_text(entry.manufacturer, width)
            elif header == 'MFGPN':
                lines = _wrap_text(entry.mfgpn, width)
            elif header == 'Datasheet':
                shortened = _shorten_url(entry.datasheet, width)
                lines = [shortened]
            elif header == 'SMD':
                lines = [entry.smd]
            elif header == 'Match_Quality':
                lines = _wrap_text(entry.match_quality, width)
            elif header == 'Priority':
                lines = [str(entry.priority)]
            elif header == 'Notes':
                lines = _wrap_text(entry.notes or '', width)
            else:
                lines = ['']
            
            cell_lines[header] = lines
            max_lines = max(max_lines, len(lines))
        
        # Print each line of the row
        for line_num in range(max_lines):
            row_line = ''
            for i, header in enumerate(headers):
                width = col_widths[header]
                # Get the content for this line, or empty string if no more lines
                content = cell_lines[header][line_num] if line_num < len(cell_lines[header]) else ''
                row_line += f"{content:<{width}}"
                if i < len(headers) - 1:
                    row_line += ' | '
            print(row_line)
    
    print()


def print_formatted_summary(file_info: List[tuple], inventory_path: Path, inventory_count: int, 
                           output_path: Path, bom_count: int, is_smd_only: bool = False, 
                           smd_excluded_count: int = 0, console_output: bool = False):
    """Print a nicely formatted summary of the BOM generation process."""
    
    # Schematic section
    if len(file_info) > 1:
        print("Hierarchical schematic set:")
        total_components = 0
        
        for count, file_path, warning in file_info:
            total_components += count
            warning_text = f" ({warning})" if warning else ""
            print(f"   {count:2d} Components      {file_path.name}{warning_text}")
        
        print("  ==============")
        print(f"   {total_components:2d} Components found in {len(file_info)} schematic files")
    else:
        count, file_path, warning = file_info[0]
        warning_text = f" ({warning})" if warning else ""
        print(f"Schematic: {count} Components from {file_path.name}{warning_text}")
    
    print()
    
    # Inventory section
    print("Inventory:")
    print(f"   {inventory_count:2d} Items       {inventory_path}")
    print()
    
    # BOM section
    if console_output:
        print("BOM:")
        smd_text = ""
        if is_smd_only:
            smd_text = " (SMD items only)"
            if smd_excluded_count > 0:
                smd_text += f" - excluded {smd_excluded_count} non-SMD entries"
        print(f"   {bom_count:2d} Entries     Console Table{smd_text}")
    else:
        print("BOM:")
        smd_text = ""
        if is_smd_only:
            smd_text = " (SMD items only)"
            if smd_excluded_count > 0:
                smd_text += f" - excluded {smd_excluded_count} non-SMD entries"
        print(f"   {bom_count:2d} Entries     {output_path}{smd_text}")


# ---- Library API (no prints/exits) -------------------------------------------------

@dataclass
class GenerateOptions:
    verbose: bool = False
    debug: bool = False
    smd_only: bool = False
    fields: Optional[List[str]] = None


def generate_bom_api(project_path: Union[str, Path], inventory_path: Union[str, Path], options: Optional[GenerateOptions] = None):
    """
    Library API to generate a BOM without printing or exiting the process.

    Returns a dict with keys:
      - file_info: List[Tuple[count:int, file_path:Path, warning:Optional[str]]]
      - inventory_count: int
      - bom_entries: List[BOMEntry]
      - smd_excluded_count: int
      - debug_diagnostics: List[dict]
      - components: List[Component]
      - available_fields: Dict[str, str]
    """
    options = options or GenerateOptions()

    proj_path = Path(project_path)
    inv_path = Path(inventory_path)

    if not inv_path.exists():
        raise FileNotFoundError(f"Inventory file not found: {inv_path}")

    file_extension = inv_path.suffix.lower()
    if file_extension not in ['.csv', '.xlsx', '.xls', '.numbers']:
        raise ValueError(f"Unsupported inventory file format: {file_extension}")
    if file_extension in ['.xlsx', '.xls'] and not EXCEL_SUPPORT:
        raise ImportError("Excel support requires openpyxl. Install with: pip install openpyxl")
    if file_extension == '.numbers' and not NUMBERS_SUPPORT:
        raise ImportError("Numbers support requires numbers-parser. Install with: pip install numbers-parser")

    # Determine schematic file(s)
    if proj_path.suffix == '.kicad_sch':
        if not proj_path.exists():
            raise FileNotFoundError(f"Schematic file not found: {proj_path}")
        schematic_path = proj_path
        search_dir = proj_path.parent
    else:
        search_dir = proj_path
        schematic_path = find_best_schematic(search_dir)
        if schematic_path is None:
            raise FileNotFoundError("No .kicad_sch file found in project directory")

    # Parse components (hierarchical aware)
    components: List[Component] = []
    file_info: List[Tuple[int, Path, Optional[str]]] = []

    processed_files = process_hierarchical_schematic(schematic_path, search_dir)
    for file_path in processed_files:
        parser_obj = KiCadParser(file_path)
        file_components = parser_obj.parse()
        components.extend(file_components)
        warning = "Warning: autosave file may be incomplete!" if file_path.name.startswith('_autosave-') else None
        file_info.append((len(file_components), file_path, warning))

    # Match inventory and generate BOM
    matcher = InventoryMatcher(inv_path)
    bom_generator = BOMGenerator(components, matcher)
    bom_entries, smd_excluded_count, debug_diagnostics = bom_generator.generate_bom(
        verbose=options.verbose, debug=options.debug, smd_only=options.smd_only
    )

    # Validate optional fields if provided
    if options.fields:
        available_fields = bom_generator.get_available_fields(components)
        invalid = [f for f in options.fields if f not in available_fields]
        if invalid:
            raise ValueError(f"Unknown fields: {', '.join(invalid)}")

    return {
        'file_info': file_info,
        'inventory_count': len(matcher.inventory),
        'bom_entries': bom_entries,
        'smd_excluded_count': smd_excluded_count,
        'debug_diagnostics': debug_diagnostics,
        'components': components,
        'available_fields': bom_generator.get_available_fields(components),
    }


# ---- CLI entrypoint -------------------------------------------------------------

# Field presets - easily extensible data structure
# All field names stored in normalized snake_case internally
# Standard BOM fields don't need qualification (reference, quantity, value, etc.)
# Inventory-specific fields are qualified with i: to avoid ambiguity
FIELD_PRESETS = {
    'standard': {
        'fields': ['reference', 'quantity', 'description', 'value', 'footprint', 'lcsc', 'datasheet', 'smd'],
        'description': 'Comprehensive set with all standard BOM fields'
    },
    'jlc': {
        'fields': ['reference', 'quantity', 'value', 'i:package', 'lcsc', 'smd'],
        'description': 'JLCPCB requirements: Reference Designator, Quantity, Value, Package, LCSC Part Number'
    },
    'minimal': {
        'fields': ['reference', 'quantity', 'value', 'lcsc'],
        'description': 'Bare minimum: reference, qty, value, and LCSC part number'
    },
    'all': {
        'fields': None,  # Special case: means "include all available fields"
        'description': 'All available fields from inventory and components'
    },
}

def _preset_fields(preset: str, include_verbose: bool, any_notes: bool) -> List[str]:
    """Build a preset field list with optional verbose/notes fields.
    
    Args:
        preset: Preset name (key from FIELD_PRESETS)
        include_verbose: Add match_quality and priority columns
        any_notes: Add notes column if there are notes in BOM
    
    Returns:
        List of field names for the preset (normalized snake_case)
    
    Raises:
        ValueError: If preset name is unknown
    """
    preset = (preset or 'standard').lower()
    
    if preset not in FIELD_PRESETS:
        raise ValueError(f"Unknown preset: {preset} (valid: {', '.join(FIELD_PRESETS.keys())})")
    
    preset_def = FIELD_PRESETS[preset]
    
    # Special case: 'all' preset is handled later when we know available fields
    if preset_def['fields'] is None:
        return []  # Placeholder, will be expanded in _parse_fields_argument
    
    result = list(preset_def['fields'])
    
    if include_verbose:
        result.append('match_quality')
    if any_notes:
        result.append('notes')
    if include_verbose:
        result.append('priority')
    
    return result


def _parse_fields_argument(fields_arg: str, available_fields: Dict[str, str], include_verbose: bool, any_notes: bool) -> List[str]:
    """
    Parse --fields argument which can contain:
    1. Preset names with + prefix: +jlc, +standard, +minimal, +all
    2. Comma-separated field names (case-insensitive, various formats)
    3. Mix of presets and fields: +jlc,CustomField1,CustomField2
    
    User input is normalized: snake_case, Title Case, CamelCase, spaces all accepted.
    Returns expanded field list (normalized snake_case) or raises ValueError for invalid fields/presets.
    """
    if not fields_arg:
        return _preset_fields('standard', include_verbose, any_notes)
    
    tokens = [t.strip() for t in fields_arg.split(',') if t.strip()]
    result = []
    known_presets = set(FIELD_PRESETS.keys())
    
    for token in tokens:
        if token.startswith('+'):
            # This is a preset expansion
            preset_name = token[1:].lower()
            if preset_name not in known_presets:
                raise ValueError(f"Unknown preset: +{preset_name} (valid: {', '.join('+' + p for p in sorted(known_presets))})")
            
            # Handle special case: +all expands to all available fields
            if preset_name == 'all':
                result.extend(sorted(available_fields.keys()))
            else:
                # Expand preset inline
                preset_fields = _preset_fields(preset_name, include_verbose, any_notes)
                result.extend(preset_fields)
        else:
            # This is a field name - normalize and validate it
            normalized_token = normalize_field_name(token)
            if normalized_token not in available_fields:
                raise ValueError(f"Unknown field: {token}. Use --list-fields to see available fields.")
            result.append(normalized_token)
    
    # Remove duplicates while preserving order
    seen = set()
    deduped = []
    for f in result:
        if f not in seen:
            seen.add(f)
            deduped.append(f)
    
    return deduped if deduped else _preset_fields('standard', include_verbose, any_notes)


def main():
    parser = argparse.ArgumentParser(description='jBOM - Generate BOM from KiCad project')
    parser.add_argument('project_path', help='Path to KiCad project directory')
    parser.add_argument('-i', '--inventory', required=True, help='Path to inventory file (.csv, .xlsx, .xls, or .numbers)')
    parser.add_argument('-o', '--output', help='Output CSV file path')
    parser.add_argument('--outdir', help='Directory for output files (used when --output is not provided)')
    parser.add_argument('-v', '--verbose', action='store_true', help='Include Match_Quality and Priority columns. Shows detailed scoring information')
    parser.add_argument('-f', '--fields', help='Field selection: use preset with + prefix (+standard, +jlc, +minimal, +all) or comma-separated field list. Mix both: +jlc,CustomField. Use --list-fields to see available fields')
    parser.add_argument('--multi-format', help='Comma-separated list of formats to emit in one run (e.g., jlc,standard)')
    parser.add_argument('--list-fields', action='store_true', help='List all available fields from inventory and component data, then exit')
    parser.add_argument('-d', '--debug', action='store_true', help='Add detailed matching information to Notes column for debugging')
    parser.add_argument('--smd', action='store_true', help='Include only SMD (Surface Mount Device) components in BOM output')
    parser.add_argument('--quiet', action='store_true', help='Suppress non-essential output (useful for CI)')
    parser.add_argument('--json-report', help='Write a JSON report with run statistics to the given path')
    
    args = parser.parse_args()
    
    project_path = Path(args.project_path)
    inventory_path = Path(args.inventory)
    
    # Validate inventory file exists and format is supported
    if not inventory_path.exists():
        print(f"Inventory file not found: {inventory_path}")
        sys.exit(1)
    
    file_extension = inventory_path.suffix.lower()
    if file_extension not in ['.csv', '.xlsx', '.xls', '.numbers']:
        print(f"Unsupported inventory file format: {file_extension}")
        print("Supported formats: .csv, .xlsx, .xls, .numbers")
        sys.exit(1)
    
    # Check for required packages
    if file_extension in ['.xlsx', '.xls'] and not EXCEL_SUPPORT:
        print(f"Excel support ({file_extension}) requires openpyxl package.")
        print("Install with: pip install openpyxl")
        sys.exit(1)
    
    if file_extension == '.numbers' and not NUMBERS_SUPPORT:
        print(f"Numbers support ({file_extension}) requires numbers-parser package.")
        print("Install with: pip install numbers-parser")
        sys.exit(1)
    
    # Detect console output options
    console_output = False
    if args.output:
        output_str = args.output.lower()
        if output_str in ['-', 'console', 'stdout']:
            console_output = True
            output_path = Path('-')  # Placeholder for console output
        else:
            output_path = Path(args.output)
    else:
        # If project_path points to a directory, we infer name from it; otherwise from parent
        out_base = project_path.name if project_path.is_dir() else project_path.stem
        base_dir = Path(args.outdir) if args.outdir else (project_path if project_path.is_dir() else project_path.parent)
        try:
            base_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        output_path = base_dir / f"{out_base}_bom.csv"

    # Collect file information for formatted output
    file_info = []
    
    # Determine schematic file(s) to process
    if project_path.suffix == '.kicad_sch':
        # Specific file provided
        if not project_path.exists():
            print(f"Schematic file not found: {project_path}")
            sys.exit(1)
        schematic_path = project_path
        search_dir = project_path.parent
    else:
        # Directory provided - find schematic files
        search_dir = project_path
        schematic_path = find_best_schematic(search_dir)
        if schematic_path is None:
            sys.exit(1)
    
    # Process schematic(s) - handle hierarchical designs
    all_components = []
    processed_files = process_hierarchical_schematic(schematic_path, search_dir)
    
    for file_path in processed_files:
        parser_obj = KiCadParser(file_path)
        file_components = parser_obj.parse()
        all_components.extend(file_components)
        
        # Check for warnings
        warning = None
        if file_path.name.startswith('_autosave-'):
            warning = "Warning: autosave file may be incomplete!"
        
        file_info.append((len(file_components), file_path, warning))
    
    components = all_components
    
    # Load inventory and match components
    matcher = InventoryMatcher(inventory_path)
    
    # Generate BOM
    bom_generator = BOMGenerator(components, matcher)
    bom_entries, smd_excluded_count, debug_diagnostics = bom_generator.generate_bom(verbose=args.verbose, debug=args.debug, smd_only=args.smd)
    
    # Handle field listing
    if args.list_fields:
        available_fields = bom_generator.get_available_fields(components)
        
        # Group fields by category
        standard_fields = {}
        inventory_fields = {}
        component_fields = {}
        
        for field, description in available_fields.items():
            if description.startswith('Inventory field:'):
                inventory_fields[field] = description
            elif description.startswith('Component property:'):
                component_fields[field] = description
            elif description.startswith('Ambiguous field:'):
                standard_fields[field] = description  # Put ambiguous fields with standard fields
            else:
                standard_fields[field] = description
        
        print("Available fields for BOM output:")
        print("=" * 60)
        
        # Standard BOM fields
        print("\nSTANDARD BOM FIELDS:")
        print("-" * 30)
        for field, description in sorted(standard_fields.items()):
            print(f"{field:<25} - {description}")
        
        # Inventory fields
        print("\nINVENTORY FIELDS:")
        print("-" * 30)
        for field, description in sorted(inventory_fields.items()):
            # Remove the "Inventory field: " prefix and show both prefixed and unprefixed versions
            clean_desc = description.replace('Inventory field: ', '')
            display_field = field.replace('I:', '', 1) if field.startswith('I:') else field
            print(f"{display_field:<25} - {clean_desc} (use: {field})")
        
        # Component properties
        if component_fields:
            print("\nCOMPONENT PROPERTIES:")
            print("-" * 30)
            for field, description in sorted(component_fields.items()):
                # Remove the "Component property: " prefix and the "C:" prefix from field name
                clean_desc = description.replace('Component property: ', '')
                display_field = field.replace('C:', '', 1) if field.startswith('C:') else field
                print(f"{display_field:<25} - {clean_desc} (use: {field})")
        
        print("\nExample usage (custom fields):")
        print(f"  python {sys.argv[0]} project.kicad_sch -i inventory.csv -f Reference,Quantity,Value,LCSC")
        print(f"\nExample usage (preset expansion):")
        print(f"  python {sys.argv[0]} project.kicad_sch -i inventory.xlsx -f +jlc")
        print(f"  python {sys.argv[0]} project.kicad_sch -i inventory.csv -f +standard")
        print(f"\nExample usage (mixed preset + custom):")
        print(f"  python {sys.argv[0]} project.kicad_sch -i inventory.csv -f +jlc,I:Tolerance,C:Voltage")
        return
    
    # Define default fields and parse custom fields if provided
    any_notes = any((e.notes or '').strip() for e in bom_entries)
    available_fields = bom_generator.get_available_fields(components)
    
    # Parse --fields argument (supports presets with + prefix or custom field lists)
    try:
        fields = _parse_fields_argument(args.fields, available_fields, args.verbose, any_notes)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)
    # JSON report (optional)
    if args.json_report:
        try:
            unmatched = sum(1 for e in bom_entries if not (e.lcsc or '').strip())
            report = {
                'project': str(project_path),
                'inventory': str(inventory_path),
                'bom_entries': len(bom_entries),
                'unmatched': unmatched,
                'smd_excluded': smd_excluded_count,
                'format': 'console' if console_output else 'csv',
                'output': '-' if console_output else str(output_path),
                'verbose': args.verbose,
                'debug': args.debug,
                'smd_only': args.smd,
            }
            with open(args.json_report, 'w', encoding='utf-8') as jf:
                json.dump(report, jf, indent=2)
        except Exception:
            pass

    if console_output:
        # Console output mode - don't write CSV file, just show table and summary
        if not args.quiet:
            print_formatted_summary(file_info, inventory_path, len(matcher.inventory), 
                                   output_path, len(bom_entries), args.smd, smd_excluded_count, console_output=True)
            # Print debug diagnostics first if enabled
            if args.debug and debug_diagnostics:
                print_debug_diagnostics(debug_diagnostics)
            # Print BOM table
            print_bom_table(bom_entries, verbose=args.verbose, include_mfg=False)
    else:
        # Normal CSV output mode
        if args.multi_format:
            formats = [f.strip().lower() for f in args.multi_format.split(',') if f.strip()]
        else:
            formats = ['standard']

        # Determine base name and directory
        out_base = (output_path.stem[:-4] if output_path.name.endswith('_bom.csv') else output_path.stem)
        out_dir = output_path.parent
        for fmt in formats:
            # If --fields was specified, use it as-is; otherwise use preset for that format
            fmt_fields = fields if args.fields else _preset_fields(fmt, args.verbose, any_notes)
            out_file = output_path if len(formats) == 1 else out_dir / f"{out_base}_bom.{fmt}.csv"
            bom_generator.write_bom_csv(bom_entries, out_file, fmt_fields)
        
        # Print formatted summary
        if not args.quiet:
            print_formatted_summary(file_info, inventory_path, len(matcher.inventory), 
                                   output_path, len(bom_entries), args.smd, smd_excluded_count, console_output=False)
            # Print debug diagnostics if debug mode is enabled and there are diagnostics
            if args.debug and debug_diagnostics:
                print_debug_diagnostics(debug_diagnostics)

    # Exit with 2 when there are unmatched entries (warning state), else 0
    try:
        unmatched_exit = sum(1 for e in bom_entries if not (e.lcsc or '').strip())
        if unmatched_exit > 0:
            sys.exit(2)
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except SystemExit as e:
        raise
    except Exception as e:
        # Hard error
        sys.exit(1)
