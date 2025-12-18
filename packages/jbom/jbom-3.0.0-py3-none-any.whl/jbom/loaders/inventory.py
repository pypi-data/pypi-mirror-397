"""
Inventory file loader for jBOM.

Handles loading inventory data from multiple file formats:
- CSV (.csv)
- Excel (.xlsx, .xls)
- Apple Numbers (.numbers)
"""

import csv
import warnings
from pathlib import Path
from typing import List, Dict

from jbom.common.types import InventoryItem, DEFAULT_PRIORITY

# Suppress specific Numbers version warning
warnings.filterwarnings(
    "ignore", message="Numbers version 14.3 not tested with this version"
)

# Optional imports for spreadsheet support
try:
    import openpyxl

    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

try:
    from numbers_parser import Document as NumbersDocument

    NUMBERS_SUPPORT = True
except ImportError:
    NUMBERS_SUPPORT = False


class InventoryLoader:
    """Loads inventory data from various file formats."""

    def __init__(self, inventory_path: Path):
        """Initialize loader with path to inventory file.

        Args:
            inventory_path: Path to inventory file (.csv, .xlsx, .xls, or .numbers)
        """
        self.inventory_path = inventory_path
        self.inventory: List[InventoryItem] = []
        self.inventory_fields: List[str] = []

    def load(self) -> tuple[List[InventoryItem], List[str]]:
        """Load inventory from supported file format.

        Returns:
            Tuple of (inventory items list, field names list)

        Raises:
            ImportError: If required package for file format is not installed
            ValueError: If file format is unsupported or file structure is invalid
        """
        file_extension = self.inventory_path.suffix.lower()

        if file_extension == ".csv":
            self._load_csv_inventory()
        elif file_extension in [".xlsx", ".xls"]:
            if not EXCEL_SUPPORT:
                raise ImportError(
                    "Excel support requires openpyxl package. Install with: pip install openpyxl"
                )
            self._load_excel_inventory()
        elif file_extension == ".numbers":
            if not NUMBERS_SUPPORT:
                raise ImportError(
                    "Numbers support requires numbers-parser package. Install with: pip install numbers-parser"
                )
            self._load_numbers_inventory()
        else:
            raise ValueError(
                f"Unsupported inventory file format: {file_extension}. Supported formats: .csv, .xlsx, .xls, .numbers"
            )

        return self.inventory, self.inventory_fields

    def _load_csv_inventory(self):
        """Load inventory from CSV file"""
        with open(self.inventory_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            headers = reader.fieldnames or []
            self._process_inventory_data(headers, rows)

    def _load_excel_inventory(self):
        """Load inventory from Excel file (.xlsx or .xls)"""
        workbook = openpyxl.load_workbook(self.inventory_path, data_only=True)
        # Use the first worksheet
        worksheet = workbook.active

        # Find the header row by looking for 'IPN' column
        header_row = None
        start_col = 1

        for row_num in range(1, min(10, worksheet.max_row + 1)):  # Check first 10 rows
            for col_num in range(
                1, min(10, worksheet.max_column + 1)
            ):  # Check first 10 columns
                cell_value = worksheet.cell(row_num, col_num).value
                if cell_value and str(cell_value).strip().upper() == "IPN":
                    header_row = row_num
                    start_col = col_num
                    break
            if header_row:
                break

        if not header_row:
            raise ValueError(
                "Could not find 'IPN' header column in Excel file. Make sure the inventory has an 'IPN' column."
            )

        # Get headers from the identified header row
        headers = []
        col_num = start_col
        while col_num <= worksheet.max_column:
            cell_value = worksheet.cell(header_row, col_num).value
            if cell_value is not None and str(cell_value).strip():
                headers.append(str(cell_value).strip())
                col_num += 1
            else:
                # Stop when we hit an empty header cell
                break

        # Get data rows
        rows = []
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            row_data = {}
            has_data = False

            for col_offset, header in enumerate(headers):
                col_num = start_col + col_offset
                if col_num <= worksheet.max_column:
                    cell_value = worksheet.cell(row_num, col_num).value
                    value_str = (
                        str(cell_value).strip() if cell_value is not None else ""
                    )
                    row_data[header] = value_str
                    if value_str:  # Check if this row has any data
                        has_data = True
                else:
                    row_data[header] = ""

            # Only add rows that have some data
            if has_data:
                rows.append(row_data)

        workbook.close()
        self._process_inventory_data(headers, rows)

    def _load_numbers_inventory(self):
        """Load inventory from Apple Numbers file"""
        doc = NumbersDocument(self.inventory_path)
        # Get the first table from the first sheet
        if not doc.sheets:
            raise ValueError("No sheets found in Numbers file")

        sheet = doc.sheets[0]
        if not sheet.tables:
            raise ValueError("No tables found in first sheet of Numbers file")

        table = sheet.tables[0]

        # Find the header row by looking for 'IPN' column (similar to Excel)
        header_row_idx = None
        start_col = None

        for row_idx in range(min(10, table.num_rows)):  # Check first 10 rows
            for col_idx in range(min(10, table.num_cols)):  # Check first 10 columns
                cell = table.cell(row_idx, col_idx)
                if cell.value and str(cell.value).strip().upper() == "IPN":
                    header_row_idx = row_idx
                    start_col = col_idx
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            raise ValueError(
                "Could not find 'IPN' header column in Numbers file. Make sure the inventory has an 'IPN' column."
            )

        # Get headers from the identified header row
        headers = []

        for col_idx in range(start_col, table.num_cols):
            cell = table.cell(header_row_idx, col_idx)
            if cell.value is not None and str(cell.value).strip():
                headers.append(str(cell.value).strip())
            else:
                # Stop when we hit an empty header cell
                break

        # Get data rows
        rows = []
        for row_idx in range(header_row_idx + 1, table.num_rows):
            row_data = {}
            has_data = False

            for col_offset, header in enumerate(headers):
                col_idx = start_col + col_offset
                if col_idx < table.num_cols:
                    cell = table.cell(row_idx, col_idx)
                    value_str = (
                        str(cell.value).strip() if cell.value is not None else ""
                    )
                    row_data[header] = value_str
                    if value_str:  # Check if this row has any data
                        has_data = True
                else:
                    row_data[header] = ""

            # Only add rows that have some data
            if has_data:
                rows.append(row_data)

        self._process_inventory_data(headers, rows)

    def _process_inventory_data(self, headers: List[str], rows: List[Dict[str, str]]):
        """Process inventory data from any source format into InventoryItem objects"""
        # Validate required headers
        required_headers = ["IPN", "Category"]
        header_upper = [h.upper() for h in headers]
        missing_headers = [
            req for req in required_headers if req.upper() not in header_upper
        ]

        if missing_headers:
            raise ValueError(
                f"Inventory file is missing required columns: {', '.join(missing_headers)}. "
                f"Found columns: {', '.join(headers)}"
            )

        # Clean up field names - replace newlines and extra whitespace
        self.inventory_fields = []
        for field in headers:
            if field:
                # Replace newlines with spaces and normalize whitespace
                clean_field = " ".join(
                    field.replace("\n", " ").replace("\r", " ").split()
                )
                self.inventory_fields.append(clean_field)

        for row in rows:
            if not row.get("IPN"):  # Skip empty rows
                continue

            # No need to parse stocking info - Priority column handles all ranking
            item = InventoryItem(
                ipn=row.get("IPN", ""),
                keywords=row.get("Keywords", ""),
                category=row.get("Category", ""),
                description=row.get("Description", ""),
                smd=row.get("SMD", ""),
                value=row.get("Value", ""),
                type=row.get("Type", ""),
                tolerance=row.get("Tolerance", ""),
                voltage=row.get("V", ""),
                amperage=row.get("A", ""),
                wattage=row.get("W", ""),
                lcsc=row.get("LCSC", ""),
                manufacturer=row.get("Manufacturer", ""),
                mfgpn=row.get("MFGPN", ""),
                datasheet=row.get("Datasheet", ""),
                package=row.get("Package", ""),
                priority=self._parse_priority(
                    row.get("Priority", str(DEFAULT_PRIORITY))
                ),
                raw_data=row,
            )
            self.inventory.append(item)

    def _parse_priority(self, priority_str: str) -> int:
        """Parse priority value from CSV, defaulting to DEFAULT_PRIORITY if invalid"""
        try:
            return (
                int(priority_str.strip()) if priority_str.strip() else DEFAULT_PRIORITY
            )
        except (ValueError, AttributeError):
            return DEFAULT_PRIORITY
