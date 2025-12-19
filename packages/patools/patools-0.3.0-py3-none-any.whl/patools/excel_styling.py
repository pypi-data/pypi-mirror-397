import json
import pydoc

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Protection, NamedStyle
import pandas

__all__ = ['get', 'apply']


def get_cell_style(cell: Cell, default_style: str):
    style = {}
    try:
        style['style'] = cell.style
    except:
        style['style'] = default_style
    style['font'] = cell.font
    style['fill'] = cell.fill
    style['border'] = cell.border
    style['alignment'] = cell.alignment
    style['protection'] = cell.protection
    style['number_format'] = cell.number_format
    return style


def style_to_json(style: dict):
    return json.dumps(
        style,
        default=lambda o: o.__dict__ if (hasattr(o, '__dict__')) else lambda x: x.__dict__ if (hasattr(x, '__dict__')) else x.__str__,
        sort_keys=True,
        ensure_ascii=False,
    )


def set_obj_attrs_by_classes_with_params(obj, params: dict, class_attributes: dict, class_subattributes: dict | None = None):
    for attr_name in params:
        found = False
        for class_name, attributes in class_attributes.items():
            if attr_name in attributes:
                found = True
                cls = pydoc.locate(class_name)  # openpyxl.styles.colors.Color
                if params[attr_name] is not None:
                    if class_subattributes is not None:
                        obj2 = cls()
                        set_obj_attrs_by_classes_with_params(obj2, params[attr_name], class_subattributes)
                        setattr(obj, attr_name, obj2)
                    else:
                        setattr(obj, attr_name, cls(**params[attr_name]))
                else:
                    setattr(obj, attr_name, cls())

        if not found:
            setattr(obj, attr_name, params[attr_name])


def get(workbook: openpyxl.Workbook) -> pandas.DataFrame:
    """
    Returns a pandas.DataFrame with styles from an openpyxl.Workbook.
    """
    if not isinstance(workbook, openpyxl.Workbook):
        raise ValueError('workbook must be an instance of openpyxl.Workbook')

    styles = []
    for worksheet in workbook.worksheets:
        for row_idx in range(worksheet.max_row):
            row_height = worksheet.row_dimensions[row_idx + 1].height
            row = {
                'Сегмент данных': 'Стили',
                'Секция таблицы': 'Ячейки',
                'Стиль строки': style_to_json({'height': row_height}),
                'Лист': worksheet.title,
                'Номер строки': row_idx + 1
            }
            column_nm = 0
            for col in worksheet.iter_cols(1, worksheet.max_column):
                column_nm += 1
                cell = col[row_idx]
                merged_cells_range = None
                if isinstance(cell, Cell):
                    for merged in worksheet.merged_cells:
                        if cell.coordinate in merged:
                            merged_cells_range = merged.coord
                            break

                column_width = worksheet.column_dimensions[get_column_letter(cell.column)].width
                try:
                    named_style = NamedStyle(cell.style).__dict__  # Стиль из предустановленных стилей документа
                except:
                    named_style = NamedStyle(workbook.named_styles[0]).__dict__  # Применение дефолтного стиля из предустановленных стилей документа
                named_style.pop('_style', None)
                named_style.pop('_wb', None)

                cell_style = get_cell_style(cell, workbook.named_styles[0])  # Стиль заданный для самой ячейки
                row[str(column_nm)] = style_to_json({
                    'merged_cells_range': merged_cells_range,
                    'column_width': column_width,
                    'named_style': named_style,
                    'cell_style': cell_style
                })
            styles.append(row)

    return pandas.DataFrame(styles)


def apply(styles: pandas.DataFrame, data: pandas.DataFrame) -> openpyxl.Workbook:
    """
    Returns an openpyxl.Workbook with styles applied on data.
    """

    def _apply(workbook: openpyxl.Workbook, dataset: pandas.DataFrame):
        dataset_columns = dataset.shape[1]
        for row_idx in range(dataset.shape[0]):
            seg_data = dataset.loc[row_idx].iloc[0]
            table_section = dataset.loc[row_idx].iloc[1]

            worksheet_name = dataset.loc[row_idx].iloc[3]
            if worksheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(worksheet_name)
            else:
                worksheet = workbook.worksheets[workbook.sheetnames.index(worksheet_name)]
            row_number = int(dataset.loc[row_idx].iloc[4])

            if seg_data == 'Таблица' and table_section == 'Таблица':
                for col_idx in range(5, dataset_columns):
                    col_number = int(dataset.columns[col_idx])
                    cell = worksheet.cell(row=row_number, column=col_number)
                    cell_value = dataset.loc[row_idx].iloc[col_idx]

                    if isinstance(cell, Cell):
                        cell.value = cell_value
            elif seg_data == 'Стили' and table_section == 'Ячейки':
                row_style_json = dataset.loc[row_idx].iloc[2]
                row_style_data = json.loads(row_style_json)

                row_height = row_style_data['height']
                if row_height is None:
                    row_height = '15.75'
                worksheet.row_dimensions[row_number].height = float(row_height)

                for col_idx in range(5, dataset_columns):
                    full_style_json = dataset.loc[row_idx].iloc[col_idx]
                    if pandas.isna(full_style_json):
                        continue
                    full_style_data = json.loads(full_style_json)

                    merged_cells_range = full_style_data['merged_cells_range']
                    if merged_cells_range is not None:
                        worksheet.merge_cells(range_string=merged_cells_range)

                    col_number = int(dataset.columns[col_idx])
                    cell = worksheet.cell(row=row_number, column=col_number)

                    column_width = full_style_data['column_width']
                    if column_width is None:
                        column_width = '13.57'
                    worksheet.column_dimensions[get_column_letter(cell.column)].width = float(column_width)

                    named_style_data = full_style_data['named_style']
                    named_style_name = named_style_data['name']
                    if named_style_name not in workbook.named_styles:
                        s = NamedStyle(name=named_style_name)
                        s.alignment = Alignment(**named_style_data['alignment'])
                        s.border = Border(**named_style_data['border'])
                        s_fill = PatternFill()
                        set_obj_attrs_by_classes_with_params(
                            s_fill, named_style_data['fill'], {'openpyxl.styles.colors.Color': ['bgColor', 'fgColor']}
                        )
                        s.fill = s_fill
                        s.font = Font(**named_style_data['font'])
                        s.builtinId = named_style_data['builtinId']
                        s.hidden = named_style_data['hidden']
                        s.number_format = named_style_data['number_format']
                        s.protection = Protection(**named_style_data['protection'])
                        workbook.add_named_style(s)

                    cell_style_data = full_style_data['cell_style']
                    cell.style = cell_style_data['style']
                    cell_font = Font()
                    set_obj_attrs_by_classes_with_params(
                        cell_font, cell_style_data['font'], {'openpyxl.styles.colors.Color': ['color']}
                    )
                    cell.font = cell_font
                    cell_fill = PatternFill()
                    set_obj_attrs_by_classes_with_params(
                        cell_fill, cell_style_data['fill'], {'openpyxl.styles.colors.Color': ['bgColor', 'fgColor']}
                    )
                    cell.fill = cell_fill
                    cell_border = Border()
                    set_obj_attrs_by_classes_with_params(
                        cell_border,
                        cell_style_data['border'],
                        {'openpyxl.styles.borders.Side': ['left', 'right', 'bottom', 'top', 'diagonal']},
                        {'openpyxl.styles.colors.Color': ['color']}
                    )
                    cell.border = cell_border
                    cell.alignment = Alignment(**cell_style_data['alignment'])
                    cell.protection = Protection(**cell_style_data['protection'])
                    cell.number_format = cell_style_data['number_format']

    workbook = openpyxl.Workbook()
    _apply(workbook, styles)
    _apply(workbook, data)

    if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook['Sheet']

    return workbook
