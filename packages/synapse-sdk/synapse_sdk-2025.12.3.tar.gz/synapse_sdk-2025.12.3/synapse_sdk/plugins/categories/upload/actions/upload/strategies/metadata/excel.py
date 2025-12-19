from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ...exceptions import ExcelParsingError, ExcelSecurityError
from ...utils import ExcelMetadataUtils, ExcelSecurityConfig
from ..base import MetadataStrategy, ValidationResult


class ExcelMetadataStrategy(MetadataStrategy):
    """Excel metadata extraction strategy."""

    def __init__(self):
        self.excel_config = ExcelSecurityConfig()
        self.excel_utils = ExcelMetadataUtils(self.excel_config)

    def extract(self, source_path: Path) -> Dict[str, Dict[str, Any]]:
        """Extract metadata from Excel file."""
        try:
            excel_stream = self._prepare_excel_file(source_path)
            workbook = load_workbook(excel_stream, read_only=True, data_only=True)
            try:
                return self._process_excel_worksheet(workbook.active)
            finally:
                workbook.close()

        except ExcelSecurityError:
            raise
        except ExcelParsingError:
            raise
        except InvalidFileException as e:
            raise ExcelParsingError(f'Invalid Excel file format: {str(e)}')
        except MemoryError:
            raise ExcelSecurityError('Excel file exceeds memory limits')
        except (OSError, IOError) as e:
            raise ExcelParsingError(f'File access error: {str(e)}')
        except Exception as e:
            # Handle ZIP file errors (Excel files are ZIP archives)
            if 'zip file' in str(e).lower():
                raise ExcelParsingError(f'Invalid Excel file format: {str(e)}')
            raise ExcelParsingError(f'Unexpected error: {str(e)}')

    def validate(self, metadata: Dict) -> ValidationResult:
        """Validate extracted metadata."""
        errors = []

        if not isinstance(metadata, dict):
            errors.append('Metadata must be a dictionary')
            return ValidationResult(valid=False, errors=errors)

        # Validate each file's metadata
        for file_name, file_metadata in metadata.items():
            if not isinstance(file_metadata, dict):
                errors.append(f"Metadata for file '{file_name}' must be a dictionary")
                continue

            # Check filename length
            if not self.excel_utils.is_valid_filename_length(file_name):
                errors.append(f"Filename '{file_name}' exceeds maximum length")

        return ValidationResult(valid=len(errors) == 0, errors=errors)

    def _validate_excel_security(self, excel_path: Path) -> None:
        """Validate Excel file security constraints."""
        file_size = excel_path.stat().st_size
        if file_size > self.excel_config.MAX_FILE_SIZE_BYTES:
            raise ExcelSecurityError(
                f'Excel file too large: {file_size} bytes (max: {self.excel_config.MAX_FILE_SIZE_BYTES})'
            )

        estimated_memory = file_size * 3
        if estimated_memory > self.excel_config.MAX_MEMORY_USAGE_BYTES:
            raise ExcelSecurityError(
                f'Excel file may consume too much memory: ~{estimated_memory} bytes '
                f'(max: {self.excel_config.MAX_MEMORY_USAGE_BYTES})'
            )

    def _prepare_excel_file(self, excel_path: Path) -> BytesIO:
        """Prepare Excel file for processing."""
        self._validate_excel_security(excel_path)
        excel_bytes = excel_path.read_bytes()
        return BytesIO(excel_bytes)

    def _process_excel_headers(self, headers: tuple) -> tuple:
        """Process Excel headers."""
        if len(headers) < 2:
            raise ExcelParsingError('Excel file must have at least 2 columns (file name and metadata)')

        # Validate first column header (filename column)
        first_header = str(headers[0]).strip().lower() if headers[0] else ''
        valid_filename_headers = ['filename', 'file_name']

        if first_header not in valid_filename_headers:
            raise ExcelParsingError(
                f'First column header must be "filename" or "file_name", got: "{headers[0]}". '
                f'Valid options: {", ".join(valid_filename_headers)}'
            )

        self._validate_excel_content(headers, 0)
        return headers

    def _process_excel_data_row(self, row: tuple, headers: tuple) -> Optional[Dict[str, Any]]:
        """Process a single Excel data row."""
        if not row[0] or str(row[0]).strip() == '':
            return None

        file_name = str(row[0]).strip()
        if not self.excel_utils.is_valid_filename_length(file_name):
            return None

        file_metadata: Dict[str, Any] = {}
        for i, value in enumerate(row[1:], start=1):
            if i < len(headers):  # Include empty strings, exclude only None values
                header_value = headers[i]
                column_name = str(header_value).strip() if header_value is not None else f'column_{i}'

                column_name = self.excel_utils.validate_and_truncate_string(
                    column_name, self.excel_config.MAX_COLUMN_NAME_LENGTH
                )

                # Convert None to empty string, otherwise convert to string
                value_str = '' if value is None else str(value)
                str_value = self.excel_utils.validate_and_truncate_string(
                    value_str, self.excel_config.MAX_METADATA_VALUE_LENGTH
                )
                file_metadata[column_name] = str_value

        return {file_name: file_metadata} if file_metadata else None

    def _process_excel_worksheet(self, worksheet) -> Dict[str, Dict[str, Any]]:
        """Process Excel worksheet."""
        if worksheet is None:
            raise ExcelParsingError('Excel file has no active worksheet')

        metadata_dict: Dict[str, Dict[str, Any]] = {}
        headers: Optional[tuple] = None
        data_row_count = 0
        validation_interval = getattr(self.excel_config, 'VALIDATION_CHECK_INTERVAL', 1000)

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            # Quick check: if row is empty or first cell is empty, skip
            if not row or not row[0] or str(row[0]).strip() == '':
                continue

            if row_idx == 0:
                headers = self._process_excel_headers(row)
                continue

            if headers is None:
                raise ExcelParsingError('Excel file missing header row')

            data_row_count += 1

            if data_row_count % validation_interval == 0:
                self._validate_excel_content(headers, data_row_count)

            row_result = self._process_excel_data_row(row, headers)
            if row_result:
                metadata_dict.update(row_result)

        self._validate_excel_content(headers or (), data_row_count)

        return metadata_dict

    def _validate_excel_content(self, headers: tuple, row_count: int) -> None:
        """Validate Excel content constraints."""
        if len(headers) > self.excel_config.MAX_COLUMNS:
            raise ExcelParsingError(f'Too many columns: {len(headers)} (max: {self.excel_config.MAX_COLUMNS})')

        if row_count > self.excel_config.MAX_ROWS:
            raise ExcelParsingError(f'Too many rows: {row_count} (max: {self.excel_config.MAX_ROWS})')
