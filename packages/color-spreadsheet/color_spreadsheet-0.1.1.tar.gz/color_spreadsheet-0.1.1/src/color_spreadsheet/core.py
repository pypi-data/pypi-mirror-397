"""Core processing logic for applying color coding to Excel workbooks."""
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from color_spreadsheet.classifier import CellType, classify_cell
from color_spreadsheet.colorways.base import Colorway
from color_spreadsheet.reporter import Statistics, SheetStatistics


def process_workbook(input_path: str, output_path: str, colorway: Colorway) -> Statistics:
    """
    Apply color coding to an Excel workbook based on cell content types.

    Args:
        input_path: Path to the input Excel file
        output_path: Path where the colored Excel file will be saved
        colorway: Colorway instance defining the color scheme

    Returns:
        Statistics object containing processing results

    Raises:
        FileNotFoundError: If input_path does not exist
        PermissionError: If unable to read input or write output
    """
    # Load workbook
    wb: Workbook = load_workbook(input_path, data_only=False)

    # Initialize statistics
    stats = Statistics()

    # Process all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_stats = SheetStatistics()

        # Iterate through all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell_type = classify_cell(cell)

                # Apply coloring based on cell type
                if cell_type == CellType.EMPTY:
                    # Skip empty cells
                    continue

                elif cell_type == CellType.HARDCODED_VALUE:
                    cell.font = colorway.get_font_for_hardcoded_value()
                    sheet_stats.blue += 1
                    stats.total_cells_processed += 1

                elif cell_type == CellType.FORMULA_SAME_SHEET:
                    cell.font = colorway.get_font_for_same_sheet_formula()
                    sheet_stats.black += 1
                    stats.total_cells_processed += 1

                elif cell_type == CellType.FORMULA_CROSS_SHEET:
                    cell.font = colorway.get_font_for_cross_sheet_formula()
                    sheet_stats.green += 1
                    stats.total_cells_processed += 1

                elif cell_type == CellType.TEXT:
                    font = colorway.get_font_for_text()
                    if font is not None:
                        cell.font = font
                    sheet_stats.text_skipped += 1
                    stats.total_cells_processed += 1

        # Store sheet statistics
        stats.sheets[sheet_name] = sheet_stats

    # Save to output file
    wb.save(output_path)

    return stats
