from __future__ import annotations

import io
import datetime


def _jsonable(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.isoformat()
    return str(v)


def convert_xlsx_bytes(data: bytes, formulas: str = "evaluated") -> dict:
    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise RuntimeError("XLSX conversion requires `openpyxl`. Install with `pip install 'kytchen[converters]'`.") from e

    data_only = str(formulas).lower() != "raw"
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=data_only, read_only=True)

    sheets: dict[str, list[list]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_jsonable(v) for v in row])
        sheets[name] = rows

    return {
        "type": "xlsx",
        "data_only": data_only,
        "sheet_names": list(wb.sheetnames),
        "sheets": sheets,
    }
