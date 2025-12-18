from __future__ import annotations

from typing import Sequence, TypeVar, TYPE_CHECKING

if TYPE_CHECKING:
    from typing import Any, Literal, Union
    from openpyxl import Workbook, _ZipFileFileProtocol
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    from openpyxl.formatting import Rule
    from openpyxl.styles import Alignment, Border, Color, PatternFill, Font

Column = TypeVar("Column", int, str)
Row = TypeVar("Row", bound=int)
Range = TypeVar("Range", bound=tuple[str,str])
Ranges = TypeVar("Ranges", list, Column, Row, Range)

StyleConfig = TypeVar("StyleConfig", bound=dict[str,dict])
RuleConfig = TypeVar("RuleConfig", bound=dict)

Width = TypeVar("Width", float, str)
Height = TypeVar("Height", float, str)
Multiple = TypeVar("Multiple", bound=str)

SINGLE_WIDTH: float = 8.43
SINGLE_HEIGHT: float = 15.0


def filter_warnings():
    import warnings
    warnings.filterwarnings("ignore", module="openpyxl.*")


def to_unique_headers(headers: list[str]) -> list[str]:
    unique = list()
    for header in headers:
        header_str, suffix = str(header), 1
        while header_str in unique:
            header_str = f"{header}_{suffix}"
            suffix += 1
        unique.append(header_str)
    return unique


def csv2json(
        io: _ZipFileFileProtocol,
        header: int = 0,
        delimiter: str = ",",
        lineterminator: str = "\r\n",
        encoding: str | None = "utf-8",
    ) -> list[dict]:
    import os
    if isinstance(io, str) and os.path.exists(io):
        with open(io, 'r', encoding=encoding) as file:
            csv2json(file, header)

    import csv
    if isinstance(io, bytes):
        from io import BytesIO, TextIOWrapper
        io = TextIOWrapper(BytesIO(io), encoding=encoding)
    rows = list(csv.reader(io, delimiter=delimiter, lineterminator=lineterminator))
    header_row = to_unique_headers(rows[header])
    return [dict(zip(header_row, row)) for row in rows[(header+1):]]


def excel2json(
        io: _ZipFileFileProtocol,
        sheet_name: str | None = None,
        header: int = 1,
        warnings: bool = True
    ) -> list[dict]:
    from openpyxl import load_workbook
    from io import BytesIO
    if not warnings:
        filter_warnings()

    wb = load_workbook(BytesIO(io) if isinstance(io, bytes) else io)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = to_unique_headers([cell.value for cell in next(ws.iter_rows(min_row=header, max_row=header))])
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=(header+1), values_only=True)]


def csv2excel(
        obj: Sequence[Sequence[Any]] | dict[str,Sequence[Sequence[Any]]],
        sheet_name: str = "Sheet1",
        header: bool = True,
        header_style: StyleConfig | Literal["yellow"] = "yellow",
        column_style: dict[Column,StyleConfig] = dict(),
        row_style: dict[Row,StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal["auto"] | None = "auto",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[tuple[Ranges,RuleConfig]] = list(),
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
    ) -> Workbook:
    from openpyxl import Workbook
    wb = Workbook()
    obj = {sheet_name: obj} if isinstance(obj, Sequence) else obj
    kwargs = dict(
        column_style=column_style, row_style=row_style,
        column_width=column_width, row_height=row_height,
        conditional_formatting=conditional_formatting, hyperlink=hyperlink,
        truncate=truncate, wrap_text=wrap_text, freeze_panes=freeze_panes)

    for index, (name, rows) in enumerate(obj.items()):
        _rows2sheet(wb, rows, index, name, header, header_style, **kwargs)
    return wb


def json2excel(
        obj: Sequence[dict] | dict[str,Sequence[dict]],
        sheet_name: str = "Sheet1",
        header: bool = True,
        header_style: StyleConfig | Literal["yellow"] = "yellow",
        column_style: dict[Column,StyleConfig] = dict(),
        row_style: dict[Row,StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal["auto"] | None = "auto",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[tuple[Ranges,RuleConfig]] = list(),
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
    ) -> Workbook:
    from openpyxl import Workbook
    wb = Workbook()
    obj = {sheet_name: obj} if isinstance(obj, Sequence) else obj
    kwargs = dict(
        column_style=column_style, row_style=row_style,
        column_width=column_width, row_height=row_height,
        conditional_formatting=conditional_formatting, hyperlink=hyperlink,
        truncate=truncate, wrap_text=wrap_text, freeze_panes=freeze_panes)

    for index, (name, rows) in enumerate(obj.items()):
        headers = list(rows[0].keys()) if rows else list()
        values = [[row.get(header, None) for header in headers] for row in rows]
        csv_rows = ([headers] if header else list()) + values
        _rows2sheet(wb, csv_rows, index, name, header, header_style, **kwargs)
    return wb


def _rows2sheet(
        wb: Workbook,
        rows: Sequence[Sequence[Any]],
        sheet_index: int,
        sheet_name: str = "Sheet1",
        header: bool = True,
        header_style: StyleConfig | Literal["yellow"] = "yellow",
        **kwargs
    ):
    if sheet_index == 0:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    if not rows:
        return

    for row in rows:
        ws.append(row)

    if not isinstance(header_style, dict):
        header_style = _yellow_header() if header_style == "yellow" else dict()

    style_sheet(ws, header, header_style, **kwargs)


def style_sheet(
        ws: Worksheet,
        header: bool = True,
        header_style: StyleConfig = dict(),
        column_style: dict[Column,StyleConfig] = dict(),
        row_style: dict[Row,StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal["auto"] | None = "auto",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[tuple[Ranges,RuleConfig]] = list(),
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
    ) -> Worksheet:
    HEADER = 1
    headers = [cell.value for cell in ws[HEADER]] if header else list()

    if truncate:
        row_height = SINGLE_HEIGHT if row_height is None else row_height
        wrap_text = True

    # STYLE CELLS BY COLUMN

    column_style = {get_column_index(column, headers): style for column, style in column_style.items()}
    column_width = _init_column_width(column_width, headers) if column_width is not None else dict()
    auto_width = {col_idx for col_idx, width in column_width.items() if isinstance(width, str)}

    for col_idx, column in enumerate(ws.columns, start=1):
        auto_width_ = (col_idx in auto_width)
        max_width = SINGLE_WIDTH

        for row_idx, cell in enumerate(column, start=1):
            text = str(x) if (x := cell.value) is not None else str()

            if auto_width_:
                max_width = max(max_width, get_cell_width(text))

            if hyperlink and text.startswith("https://"):
                cell.hyperlink = text
                cell.font = _font(color="#0000FF", underline="single")

            if wrap_text:
                cell.alignment = _alignment(wrap_text=True)

            if header and (row_idx == HEADER):
                if header_style:
                    style_cell(cell, **header_style)
            elif col_idx in column_style:
                style_cell(cell, **column_style[col_idx])
            elif row_idx in row_style:
                style_cell(cell, **row_style[row_idx])

        # CHANGE COLUMN WIDTH

        width = min(max_width + 2., 25.) if auto_width_ else column_width.get(col_idx)
        if isinstance(width, float):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # CHANGE ROW HEIGHT

    row_height = _init_row_height(row_height) if row_height is not None else dict()

    if isinstance(row_height, dict):
        for row_idx, height in row_height.items():
            ws.row_dimensions[row_idx].height = height
    elif row_height is not None:
        for row_idx in range(1, ws.max_row+1):
            ws.row_dimensions[row_idx].height = SINGLE_HEIGHT

    # CONDITIONAL FORMATTING

    for ranges, config in conditional_formatting:
        range_string = get_range(ranges,
            min_column='A', max_column=get_column_letter(ws.max_column),
            min_row=(1+int(header)), max_row=ws.max_row, headers=headers)
        if range_string:
            ws.conditional_formatting.add(range_string, conditional_rule(**config))

    if freeze_panes:
        ws.freeze_panes = freeze_panes


def style_cell(
        cell: Cell,
        align: dict | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        number_format: str | None = None,
        hyperlink: str | None = None,
        **kwargs
    ):
    if align:
        cell.alignment = _alignment(**align)
    if border:
        cell.border = _border(**border)
    if fill:
        cell.fill = _fill(**fill)
    if font:
        cell.font = _font(**font)
    if number_format is not None:
        cell.number_format = number_format
    if hyperlink is not None:
        cell.hyperlink = hyperlink


def conditional_rule(
        operator: Literal[
            "endsWith", "containsText", "beginsWith", "lessThan", "notBetween", "lessThanOrEqual",
            "notEqual", "notContains", "between", "equal", "greaterThanOrEqual", "greaterThan"],
        formula: Sequence,
        stop_if_true: bool | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        **kwargs
    ) -> Rule:
    from openpyxl.formatting.rule import CellIsRule
    styles = dict()
    if border:
        styles["border"] = _border(**border)
    if fill:
        styles["fill"] = _fill(**fill)
    if font:
        styles["font"] = _font(**font)
    return CellIsRule(operator=operator, formula=formula, stopIfTrue=stop_if_true, **styles)


def get_range(
        ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
        min_column: str,
        max_column: str,
        min_row: int,
        max_row: int,
        headers: list[str] = list(),
    ) -> str:
    if isinstance(ranges, list):
        args = (min_column, max_column, min_row, max_row, headers)
        return ' '.join([string for range_ in ranges if (string := get_range(range_, *args))])
    elif isinstance(ranges, str):
        column = get_column_letter(ranges, headers)
        return f"{column}{min_row}:{column}{max_row}" if column is not None else str()
    elif isinstance(ranges, int):
        return f"{min_column}{ranges}:{max_column}{ranges}"
    elif isinstance(ranges, tuple) and (len(ranges) == 2):
        return f"{ranges[0]}:{ranges[1]}"
    else:
        return str()


def get_column_index(column: Column, headers: list[str] = list()) -> int | None:
    if isinstance(column, int):
        return column
    elif isinstance(column, str):
        if column.startswith('!'):
            exclude, column = -1, column[1:]
        else:
            exclude = 1
        if column in headers:
            return (headers.index(column) + 1) * exclude
    return None


def get_column_letter(column: Column, headers: list[str] = list()) -> str | None:
    from openpyxl.utils import get_column_letter as get_letter
    col_idx = get_column_index(column, headers)
    return get_letter(col_idx) if isinstance(col_idx, int) else None


def get_cell_width(value: str) -> float:
    try:
        # 한글: 1.8배, 공백: 1.2배, 영문/숫자: 1배
        return sum(1.8 if ord(c) > 12799 else 1.2 if c.isspace() else 1. for c in value)
    except:
        return 0.


def _init_column_width(
        column_width: dict[Column,Width] | float | Multiple | Literal["auto"],
        headers: list[str],
    ) -> dict[int, Union[float,Literal["auto"]]]:

    def _set_width(value: Width) -> float | Literal["auto"]:
        if isinstance(value, str):
            if value == "auto":
                return "auto"
            elif value.endswith('x'):
                value = SINGLE_WIDTH * float(value[:-1])
        return float(value) if isinstance(value, (float,int)) and value > 0. else None

    if isinstance(column_width, dict):
        return {col_idx: width for column, value in column_width.items()
                if ((col_idx := get_column_index(column, headers)) is not None)
                    and ((width := _set_width(value)) is not None)}
    else:
        value = _set_width(column_width)
        if value is not None:
            return {col_idx: value for column in headers
                    if (col_idx := get_column_index(column, headers)) is not None}
        else:
            return dict()


def _init_row_height(
        row_height: dict[Row,Height] | float | Multiple | Literal["single"],
    ) -> dict[int, float] | float | None:

    def _set_height(value: Width) -> float:
        if isinstance(value, str):
            if value == "single":
                return SINGLE_HEIGHT
            elif value.endswith('x'):
                value = SINGLE_HEIGHT * float(value[:-1])
        return float(value) if isinstance(value, (float,int)) and value > 0. else None

    if isinstance(row_height, dict):
        return {row_idx: height for row_idx, value in row_height.items()
                if (height := _set_height(value)) is not None}
    else:
        return _set_height(row_height)


def _alignment(**kwargs) -> Alignment:
    from openpyxl.styles import Alignment
    return Alignment(**kwargs)


def _border(**kwargs: dict) -> Border:
    from openpyxl.styles import Border, Side
    def side(color: str | None = None, **kwargs) -> Side:
        return Side(color=(_color(color) if color is not None else None), **kwargs)
    return Border(**{property: side(**config) for property, config in kwargs.items()})


def _fill(color: str | None = None, **kwargs) -> PatternFill:
    from openpyxl.styles import PatternFill
    for property, value in kwargs.items():
        if property in {"fgColor","bgColor","start_color","end_color"}:
            kwargs[property] = _color(value)
    if color is not None:
        color = _color(color)
        kwargs.update(start_color=color, end_color=color)
    return PatternFill(**kwargs)


def _font(color: str | None = None, **kwargs) -> Font:
    from openpyxl.styles import Font
    return Font(color=(_color(color) if color is not None else None), **kwargs)


def _color(rgb: Any, alpha: str = "FF") -> Color:
    from openpyxl.styles import Color
    if isinstance(rgb, str):
        return Color((alpha + rgb[1:]) if rgb.startswith('#') else rgb)
    elif isinstance(rgb, dict):
        return Color(**rgb)
    elif isinstance(rgb, Color):
        return rgb
    else:
        return None


def _yellow_header() -> StyleConfig:
    return {
        "align": {"horizontal": "center"},
        "fill": {"color": "#FFFF00", "fill_type": "solid"},
        "font": {"color": "#000000", "bold": True},
    }
