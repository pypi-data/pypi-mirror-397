#  This work is based on original code developed and copyrighted by TNO 2023.
#  Subsequent contributions are licensed to you by the developers of such code and are
#  made available to the Project under one or several contributor license agreements.
#
#  This work is licensed to you under the Apache License, Version 2.0.
#  You may obtain a copy of the license at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Contributors:
#      TNO         - Initial implementation
#  Manager:
#      TNO
import locale

import openpyxl
import pytz

from esdl.profiles.profilemanager import ProfileManager, DuplicateValueInProfileException, ProfileType
from openpyxl import Workbook

from esdl.utils.datetime_utils import parse_date


class ExcelProfileManager(ProfileManager):
    """
    ExcelProfileManager: manages profile data that can be loaded from and saved to excel.

    ExcelProfileManager is a subclass of ProfileManager, so it also provides functionality to convert from/to
    different ESDL profiles and to load/save to CSV
    """
    def __init__(self, source_profile=None):
        """
        Constructor of the ExcelProfileManager

        :param source_profile: the source profile of which the data is copied into this profiles manager instance
        """
        super().__init__()

        if source_profile:
            self.convert(source_profile)

    def load_excel(self, file_path, sheet_name: str = None):
        """
        Reads profile data from an Excel file.

        :param file_path: the path to the Excel file
        :param sheet_name: the name of the sheet that contains the data. By default, the first sheet is used
        """
        self.clear_profile()
        self.profile_type = ProfileType.EXCEL

        wb = openpyxl.load_workbook(file_path)
        if sheet_name:
            wb.active = sheet_name
        sheet = wb.active

        self.profile_header = [cell.value for cell in sheet[1]]
        self.num_profile_items = 0

        previous_datetime = None
        for row in sheet.iter_rows(min_row=2):
            try:
                dt = parse_date(row[0].value)
            except ValueError:  # ValueError: row in Excel doesn't start with recognizable datetime value, ignore row
                continue

            try:
                aware_dt = pytz.utc.localize(dt)  # Assume timezone is UTC if no TZ was given
            except ValueError:  # ValueError: No naive datetime (tzinfo is already set)
                aware_dt = dt
            row[0].value = aware_dt
            for elem_idx in range(1, len(row)):
                try:
                    row[elem_idx].value = float(row[elem_idx].value)      # Assume float values
                except Exception as e:
                    raise Exception(f"Cannot parse profile value in CSV ({row[elem_idx].value})")

            if previous_datetime:
                if previous_datetime == aware_dt:
                    raise DuplicateValueInProfileException(
                        "CSV contains duplicate datetimes ({}). Check timezone and daylight saving".
                        format(aware_dt.strftime('%Y-%m-%dT%H:%M:%S%z'))
                    )
            previous_datetime = aware_dt

            if self.start_datetime is None:
                self.start_datetime = aware_dt

            self.profile_data_list.append([cell.value for cell in row])
            self.num_profile_items += 1

        self.end_datetime = self.profile_data_list[-1][0]

    def save_excel(self, file_path, sheet_name: str = "Sheet1"):
        """
        Saves profile data to an Excel file.

        :param file_path: the path to the Excel file
        :param sheet_name: the name of the sheet that should be created to store the data (default is "Sheet1")
        """

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        ws.append(self.profile_header)
        for row in self.profile_data_list:
            out_row = row[:]
            out_row[0] = out_row[0].strftime('%Y-%m-%dT%H:%M:%SZ')
            ws.append(out_row)

        wb.save(file_path)
