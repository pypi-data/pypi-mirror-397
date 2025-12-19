import logging


from typing import Dict, List, Optional
from datetime import datetime


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


def autosize_columns(sheet) -> None:
    """Auto-adjust column widths based on content length."""
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 80)


def apply_frozen_and_filter(sheet) -> None:
    """Freeze header row and enable auto-filter."""
    sheet.freeze_panes = "A2"
    # Auto-filter over the full used range
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def write_sheet_by_date(
    wb: Workbook,
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
    colors: Dict[str, str],
) -> None:
    """
    Sheet 1: timeline_by_date
    Sorted by commit datetime ascending.
    Columns:
        Date, DateIndex, TopoIndex, CommitHash, Message, Author, PR_Number, <branches...>
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    headers = [
        "Date",
        "DateIndex",
        "TopoIndex",
        "CommitHash",
        "Message",
        "Author",
        "PR_Number",
    ] + branches
    sheet.append(headers)

    # Sort commits by chronological date
    chronological = sorted(
        all_commits.items(),
        key=lambda kv: (kv[1]["datetime"], kv[0]),
    )

    for _, meta in chronological:
        # some commits might not have topo_index/date_index if something is odd; guard
        date_index = meta.get("date_index", "")
        topo_index = meta.get("topo_index", "")

        base_row = [
            meta["date"],
            date_index,
            topo_index,
            meta["hash"],
            meta["message"],
            meta["author"],
            meta["pr"],
        ]

        # Branch cells + color logic
        row_values = list(base_row)
        missing_color = PatternFill(
            start_color=colors["missing"],
            end_color=colors["missing"],
            fill_type="solid",
        )
        all_present_color = PatternFill(
            start_color=colors["all_present"],
            end_color=colors["all_present"],
            fill_type="solid",
        )

        # branches start at column index:
        base_offset = len(base_row)  # 0-based for Python, 1-based in Excel
        all_present = len(meta["branches"]) == len(branches)

        for branch in branches:
            if branch in meta["branches"]:
                row_values.append("FOUND")
            else:
                row_values.append("MISSING")

        sheet.append(row_values)
        row_idx = sheet.max_row

        # Apply colors to branch cells
        for i, branch in enumerate(branches):
            col_idx = base_offset + i + 1  # 1-based
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value == "MISSING":
                cell.fill = missing_color
            elif all_present:
                cell.fill = all_present_color

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


def write_sheet_by_topology(
    wb: Workbook,
    topo_commits: List[str],
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
    colors: Dict[str, str],
) -> None:
    """
    Sheet 2: timeline_by_topology
    Ordered by topo_index (Git ancestry).
    Columns:
        TopoIndex, Date, DateIndex, CommitHash, Message, Author, PR_Number, <branches...>

    out_of_order (date earlier than previous topo commit):
        Date cell highlighted with out_of_order color.
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    headers = [
        "TopoIndex",
        "Date",
        "DateIndex",
        "CommitHash",
        "Message",
        "Author",
        "PR_Number",
    ] + branches
    sheet.append(headers)

    missing_fill = PatternFill(
        start_color=colors["missing"],
        end_color=colors["missing"],
        fill_type="solid",
    )
    all_present_fill = PatternFill(
        start_color=colors["all_present"],
        end_color=colors["all_present"],
        fill_type="solid",
    )
    out_of_order_fill = PatternFill(
        start_color=colors["out_of_order"],
        end_color=colors["out_of_order"],
        fill_type="solid",
    )

    for commit_hash in topo_commits:
        if commit_hash not in all_commits:
            continue
        meta = all_commits[commit_hash]

        topo_index = meta.get("topo_index", "")
        date_index = meta.get("date_index", "")

        base_row = [
            topo_index,
            meta["date"],
            date_index,
            meta["hash"],
            meta["message"],
            meta["author"],
            meta["pr"],
        ]

        row_values = list(base_row)
        base_offset = len(base_row)
        all_present = len(meta["branches"]) == len(branches)

        for branch in branches:
            if branch in meta["branches"]:
                row_values.append("FOUND")
            else:
                row_values.append("MISSING")

        sheet.append(row_values)
        row_idx = sheet.max_row

        # Apply branch cell colors
        for i, branch in enumerate(branches):
            col_idx = base_offset + i + 1
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value == "MISSING":
                cell.fill = missing_fill
            elif all_present:
                cell.fill = all_present_fill

        # Highlight out-of-order commits
        if meta.get("out_of_order"):
            # Date column is 2 (1-based)
            date_cell = sheet.cell(row=row_idx, column=2)
            date_cell.fill = out_of_order_fill

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


def write_sheet_hybrid(
    wb: Workbook,
    topo_commits: List[str],
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
    colors: Dict[str, str],
) -> None:
    """
    Sheet 3: timeline_hybrid

    Also topo-order, but shows both topo_index and date_index,
    with explicit out-of-order flag and days-since-previous-topo.

    Columns:
        TopoIndex, Date, DateIndex, CommitHash, Message, Author, PR_Number,
        OutOfOrder, DaysSincePrevTopo, <branches...>
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    headers = [
        "TopoIndex",
        "Date",
        "DateIndex",
        "CommitHash",
        "Message",
        "Author",
        "PR_Number",
        "OutOfOrder",
        "DaysSincePrevTopo",
    ] + branches
    sheet.append(headers)

    missing_fill = PatternFill(
        start_color=colors["missing"],
        end_color=colors["missing"],
        fill_type="solid",
    )
    all_present_fill = PatternFill(
        start_color=colors["all_present"],
        end_color=colors["all_present"],
        fill_type="solid",
    )
    out_of_order_fill = PatternFill(
        start_color=colors["out_of_order"],
        end_color=colors["out_of_order"],
        fill_type="solid",
    )

    prev_dt: Optional[datetime] = None

    for commit_hash in topo_commits:
        if commit_hash not in all_commits:
            continue
        meta = all_commits[commit_hash]

        topo_index = meta.get("topo_index", "")
        date_index = meta.get("date_index", "")

        dt = meta["datetime"]
        if prev_dt is None:
            days_since_prev = ""
        else:
            days_since_prev = (dt.date() - prev_dt.date()).days
        prev_dt = dt

        out_of_order_flag = "YES" if meta.get("out_of_order") else "NO"

        base_row = [
            topo_index,
            meta["date"],
            date_index,
            meta["hash"],
            meta["message"],
            meta["author"],
            meta["pr"],
            out_of_order_flag,
            days_since_prev,
        ]

        row_values = list(base_row)
        base_offset = len(base_row)
        all_present = len(meta["branches"]) == len(branches)

        for branch in branches:
            if branch in meta["branches"]:
                row_values.append("FOUND")
            else:
                row_values.append("MISSING")

        sheet.append(row_values)
        row_idx = sheet.max_row

        # Branch cell colors
        for i, branch in enumerate(branches):
            col_idx = base_offset + i + 1
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value == "MISSING":
                cell.fill = missing_fill
            elif all_present:
                cell.fill = all_present_fill

        # Highlight OutOfOrder cell if YES
        if meta.get("out_of_order"):
            # OutOfOrder column index is 8 (1-based)
            cell = sheet.cell(row=row_idx, column=8)
            cell.fill = out_of_order_fill

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)


def write_sheet_analytics(
    wb: Workbook,
    all_commits: Dict[str, dict],
    branches: List[str],
    sheet_name: str,
) -> None:
    """
    Sheet 4: analytics_summary

    Simple summary metrics:
    - Total unique commits
    - Commits present in all branches
    - Commits with out_of_order flag
    - For each branch: total, unique-to-branch, missing-from-branch
    """
    logger.info("Writing sheet: %s", sheet_name)
    sheet = wb.create_sheet(title=sheet_name)

    # Basic metrics
    total_unique = len(all_commits)
    all_branches_set = set(branches)

    present_in_all = sum(
        1 for meta in all_commits.values()
        if meta["branches"] == all_branches_set
    )
    out_of_order_count = sum(
        1 for meta in all_commits.values()
        if meta.get("out_of_order")
    )

    sheet.append(["Metric", "Value"])
    sheet.append(["Total unique commits across all branches", total_unique])
    sheet.append(["Commits present in all branches", present_in_all])
    sheet.append(["Commits with out-of-order dates (vs topology)", out_of_order_count])
    sheet.append([])

    # Per-branch table
    sheet.append(["Branch", "TotalCommitsInBranch", "UniqueToBranch", "MissingFromBranch"])

    for branch in branches:
        total_in_branch = sum(
            1 for meta in all_commits.values()
            if branch in meta["branches"]
        )
        unique_to_branch = sum(
            1 for meta in all_commits.values()
            if meta["branches"] == {branch}
        )
        missing_from_branch = total_unique - total_in_branch

        sheet.append([branch, total_in_branch, unique_to_branch, missing_from_branch])

    autosize_columns(sheet)
    apply_frozen_and_filter(sheet)

