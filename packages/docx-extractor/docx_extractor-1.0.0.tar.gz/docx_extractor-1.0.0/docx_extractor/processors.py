import openpyxl
from pathlib import Path


def process_excel(path):
    """Extract Excel sheets as list of rows."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if row:  # skip empty rows
                    rows.append(row)
            if rows:  # skip empty sheets
                sheets[sheet_name] = rows
        return sheets if sheets else None
    except:
        return None


def process_text(path):
    """Read text file, skip if empty."""
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read().strip()
            return content if content else None
    except:
        return None
