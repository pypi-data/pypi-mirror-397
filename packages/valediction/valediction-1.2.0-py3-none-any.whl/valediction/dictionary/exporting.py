from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.worksheet import Worksheet

from valediction.dictionary.exporting_helpers import (
    CalculatedColInfo,
    CFRuleInfo,
    DVRuleInfo,
    _build_row_list_from_mapping,
    _collect_conditional_formats,
    _collect_data_validations,
    _collect_table_formulas,
    _extend_cf_for_new_row,
    _extend_dv_for_new_row,
    _find_label_cell,
    _first_blank_data_row,
    _table_column_index_map,
)
from valediction.dictionary.integrity import (
    DD_COLUMN_MAP,
    DD_TABLE_MAP,
)
from valediction.dictionary.model import Dictionary
from valediction.exceptions import (
    DataDictionaryError,
    DataDictionaryExportError,
    DataDictionaryImportError,
)
from valediction.integrity import get_config


class Exporter:
    def __init__(
        self,
        dictionary: Dictionary | None = None,
        debug: bool = False,
        template_path: str | Path | None = None,
    ):
        # Settings
        self.debug: bool = debug
        self.debug_report: list[str] = []
        self.template_path: Path = (
            Path(template_path)
            if template_path is not None
            else Path(get_config().template_data_dictionary_path)
        )

        # Workbook
        self.workbook: Workbook | None = None

        # Dictionary
        self.dictionary: Dictionary | None = None

        # Lookups
        self.sheets: dict[str, Worksheet] = {}
        self.tables: dict[str, ExcelTable] = {}
        self.table_column_maps: dict[str, dict[str, int]] = {}

        # Formatting
        self.conditional_formats: dict[str, list[CFRuleInfo]] = {}
        self.data_validations: dict[str, list[DVRuleInfo]] = {}
        self.calculated_columns: dict[str, list[CalculatedColInfo]] = {}

        # Setup
        if dictionary:
            self.load_dictionary(dictionary)
        else:
            self.__say(
                "Exporter instantiated without dictionary - remember to load_dictionary()"
            )
        self.load_template()

    # Magic #
    def __repr__(self):
        return f"Exporter(debug={self.debug})"

    def __say(
        self,
        *values: object,
        sep: str | None = " ",
        end: str | None = "\n",
    ) -> None:
        msg = sep.join(map(str, values)) + (end or "")
        self.debug_report.append(msg)
        if self.debug:
            print("Exporter:", *values, sep=sep, end=end)

    # Main Functions #

    # High Helpers #
    def load_dictionary(self, dictionary: Dictionary):
        if not isinstance(dictionary, Dictionary):
            raise DataDictionaryError("Dictionary must be a Dictionary object")
        self.dictionary = dictionary

    def load_template(self) -> None:
        supported = {".xltx", ".xlsx"}
        if not self.template_path.exists():
            raise DataDictionaryImportError(f"Template not found: {self.template_path}")

        if self.template_path.suffix.lower() not in supported:
            raise DataDictionaryImportError(
                f"Unexpected template extension: {self.template_path.suffix}; supported: {supported}"
            )

        self.workbook = load_workbook(
            self.template_path, data_only=False, keep_vba=False
        )

        self.workbook.template = False
        self.__say(f"Loaded workbook from {self.template_path}")
        self.load_sheets_and_tables()
        self._validate_required_columns()
        self.inspect_template_features()

    def load_sheets_and_tables(self) -> None:
        if self.workbook is None:
            raise DataDictionaryImportError("Workbook not loaded.")

        for sheet_key, table_name in DD_TABLE_MAP.items():
            worksheet = self._get_worksheet(sheet_key)
            self.sheets[sheet_key.lower()] = worksheet
            attr_name = f"{sheet_key.lower()}_ws"
            setattr(self, attr_name, worksheet)

            if table_name:
                table = self._get_table(sheet_key, table_name)
                self.tables[table_name.lower()] = table

        self.__say("Required sheets and tables found and loaded.")

    def inspect_template_features(self) -> None:
        self.__say(
            "Inspecting Conditional Formatting [CF], Data Validations [DV], Calculated Columns [CC]:"
        )
        # 1) Conditional formatting per sheet
        for key, worksheet in self.sheets.items():
            rules = list(_collect_conditional_formats(worksheet))
            self.conditional_formats[key] = rules
            self.__say(f"[CF] {worksheet.title}: {len(rules)} rules")

        # 2) Data validation per sheet
        for key, worksheet in self.sheets.items():
            dv_rules = list(_collect_data_validations(worksheet))
            self.data_validations[key] = dv_rules
            self.__say(f"[DV] {worksheet.title}: {len(dv_rules)} rules")

        # 3) Calculated columns per table
        for tkey, table in self.tables.items():
            worksheet = self._worksheet_of_table(table)
            cols = _collect_table_formulas(table)
            self.calculated_columns[tkey] = cols
            self.__say(
                f"[CC] {self._table_name(table)}: {len(cols)} calculated columns"
            )

        self.__say("Inspection complete")

    def add_row(
        self, table: str, data: Optional[List] = None, quiet: bool = False
    ) -> int:
        data = [] if data is None else list(data)
        tkey = table.lower()
        tbl = self.tables.get(tkey)
        if tbl is None:
            raise DataDictionaryImportError(f"Table '{table}' not loaded.")

        ws = self._worksheet_of_table(tbl)
        min_c, min_r, max_c, max_r = range_boundaries(tbl.ref)
        ncols = max_c - min_c + 1

        # Column map (lower header -> 0-based index within table)
        colmap = self.table_column_maps.get(tkey)
        if not colmap:
            colmap = _table_column_index_map(tbl)
            self.table_column_maps[tkey] = colmap

        # Calculated/formula columns from metadata (lower header -> '=...')
        formulas: Dict[str, str] = _collect_table_formulas(tbl)

        # Data length sanity
        if len(data) > ncols:
            raise DataDictionaryError(
                f"Too many values for '{self._table_name(tbl)}': {len(data)} provided, {ncols} columns."
            )
        if len(data) < ncols:
            data += [None] * (ncols - len(data))

        # Decide target row: reuse first blank data row, else append
        reuse_row = _first_blank_data_row(ws, tbl, colmap, formulas)
        append = reuse_row is None
        target_row = (max_r + 1) if append else reuse_row

        # Write row; ignore user values for formula columns
        idx_to_header = {v: k for k, v in colmap.items()}
        for j_idx in range(1, ncols + 1):
            abs_col = min_c + j_idx - 1
            cell = ws.cell(target_row, abs_col)
            header_lower = idx_to_header.get(j_idx, "")
            if header_lower in formulas and isinstance(formulas[header_lower], str):
                cell.value = formulas[header_lower]
            else:
                cell.value = data[j_idx - 1]

        # Only extend table + CF/DV if we actually appended
        if append:
            old_bottom = max_r
            new_ref = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{target_row}"
            tbl.ref = new_ref
            if getattr(tbl, "autoFilter", None) is not None:
                tbl.autoFilter.ref = new_ref

            _extend_cf_for_new_row(
                ws, (min_c, max_c), old_bottom=old_bottom, new_row=target_row
            )
            _extend_dv_for_new_row(ws, (min_c, max_c), new_row=target_row)

        if not quiet:
            self.__say(
                f"{'Appended' if append else 'Reused'} row {target_row} in '{self._table_name(tbl)}' (width={ncols})."
            )
        return target_row

    def export(
        self,
        directory: Path | str,
        filename: str | None = None,
        overwrite: bool = False,
    ) -> Path:
        directory = Path(directory)
        filename = f"{filename}.xlsx" if filename else "PROJECT - Data Dictionary.xlsx"
        out_path = directory / filename

        # Prepare destination
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Overwrite check
        if out_path.exists() and not overwrite:
            raise DataDictionaryExportError(
                f"File exists and overwrite=False: {out_path}"
            )

        # Save
        try:
            self.workbook.save(out_path)
        except PermissionError as e:
            raise DataDictionaryExportError(
                f"Cannot write '{out_path}'. Is the file open? ({e})"
            ) from e

        self.__say(f"Exported workbook → {out_path}")
        # optional: remember last export location
        self.last_export_path = out_path
        self.workbook.close()
        return out_path

    def add_rows(
        self, table: str, rows: Iterable[Mapping[str, Any] | List[Any]]
    ) -> List[int]:
        tkey = table.lower()
        tbl = self.tables.get(tkey)
        if tbl is None:
            raise DataDictionaryImportError(f"Table '{table}' not loaded.")

        colmap = self.table_column_maps.get(tkey) or _table_column_index_map(tbl)
        self.table_column_maps[tkey] = colmap

        formulas: Dict[str, str] = _collect_table_formulas(tbl)

        written_rows: List[int] = []
        for row in rows:
            if isinstance(row, Mapping):
                data_list = _build_row_list_from_mapping(tbl, colmap, formulas, row)
            else:
                data_list = list(row)
            written_rows.append(self.add_row(table, data_list, quiet=True))

        self.__say(f"Wrote {len(written_rows)} row(s) to '{self._table_name(tbl)}'")
        return written_rows

    def write_details(self) -> dict[str, str]:
        if self.dictionary is None:
            raise DataDictionaryError(
                "No Dictionary loaded. Call load_dictionary() first."
            )

        ws = self.sheets.get("details") or self._get_worksheet("details")

        plan: list[tuple[str, object]] = [
            ("Name", self.dictionary.name),
            ("Organisation(s)", self.dictionary.organisations),
            ("Version", self.dictionary.version),
            ("Version Notes", self.dictionary.version_notes),
            ("Inclusion Criteria", self.dictionary.inclusion_criteria),
            ("Exclusion Criteria", self.dictionary.exclusion_criteria),
        ]

        written: dict[str, str] = {}
        for label, value in plan:
            if value is None:
                continue  # leave whatever the template has
            r, c = _find_label_cell(ws, label)
            dest_col = c + 1  # write in the next cell to the right
            ws.cell(r, dest_col).value = value
            addr = f"{get_column_letter(dest_col)}{r}"
            written[label] = addr
            self.__say(f"[Details] {label} → {addr}")

        return written

    def write_tables(self) -> List[int]:
        rows = []
        for table in self.dictionary:
            rows.append(
                {
                    "table": table.name,
                    "description": table.description,
                }
            )
        self.__say(f"Writing {len(rows)} row(s) to Tables")
        return self.add_rows("tables", rows)

    def write_columns(self) -> List[int]:
        rows = []
        for table in self.dictionary:
            for column in table:
                rows.append(
                    {
                        "table": table.name,
                        "column": column.name,
                        "order": column.order,
                        "data type": str(
                            column.data_type
                        ),  # DataType -> 'Text', 'Integer', …
                        "length": column.length
                        if column.data_type.allows_length()
                        else None,
                        "vocabularies": column.vocabulary,
                        "enumerations": "Y" if column.enumerations else None,
                        "primary key": column.primary_key,
                        "foreign key target": column.foreign_key,
                        "column description": column.description,
                        # 'checks' is calculated by the template
                    }
                )
        self.__say(f"Writing {len(rows)} row(s) to Columns")
        return self.add_rows("columns", rows)

    def write_enumerations(self) -> List[int]:
        rows = []
        for table in self.dictionary:
            for column in table:
                if not column.enumerations:
                    continue
                for code, name in column.enumerations.items():
                    rows.append(
                        {
                            "table": table.name,
                            "column": column.name,
                            "code": code,
                            "name": name,
                            # 'checks' is calculated by the template
                        }
                    )
        self.__say(f"Writing {len(rows)} row(s) to Enumerations")
        return self.add_rows("enumerations", rows)

    # Low Helpers #
    def _get_worksheet(self, sheet_name: str) -> Worksheet:
        """Return a worksheet by name (case-insensitive) and cache in self.sheets."""
        target = sheet_name.lower()
        matches = [ws for ws in self.workbook.worksheets if ws.title.lower() == target]
        if not matches:
            raise DataDictionaryImportError(f"Required sheet not found: '{sheet_name}'")
        if len(matches) > 1:
            titles = ", ".join(ws.title for ws in matches)
            raise DataDictionaryImportError(
                f"Duplicate sheets for '{sheet_name}' (case-insensitive): {titles}"
            )
        ws = matches[0]
        self.sheets[target] = ws
        return ws

    def _table_name(self, table_name: ExcelTable) -> str:
        """Return canonical table name from workbook."""
        return (
            getattr(table_name, "displayName", None)
            or getattr(table_name, "name", None)
            or ""
        ).strip()

    def _get_table(self, sheet_name: str, table_name: str) -> ExcelTable:
        """Return a unique table and cache in self.tables."""
        ws_expected = self._get_worksheet(sheet_name)
        target = table_name.lower()

        matches = []
        for worksheet in self.workbook.worksheets:
            for tbl in getattr(worksheet, "tables", {}).values():
                if self._table_name(tbl).lower() == target:
                    matches.append((worksheet, tbl))

        if not matches:
            raise DataDictionaryImportError(f"Required table not found: '{table_name}'")
        if len(matches) > 1:
            locs = ", ".join(
                f"{self._table_name(table)}@{worksheet.title}"
                for worksheet, table in matches
            )
            raise DataDictionaryImportError(
                f"Table name '{table_name}' is not unique across workbook: {locs}"
            )

        ws_found, tbl = matches[0]
        if ws_found is not ws_expected:
            raise DataDictionaryImportError(
                f"Table '{self._table_name(tbl)}' is on sheet '{ws_found.title}', "
                f"expected on '{ws_expected.title}'"
            )

        self.tables[target] = tbl
        return tbl

    def _worksheet_of_table(self, tbl: ExcelTable) -> Worksheet:
        """Find and return the worksheet that owns this table."""
        for ws in self.workbook.worksheets:
            for t in getattr(ws, "tables", {}).values():
                if t is tbl:
                    return ws
        raise DataDictionaryImportError("Table does not belong to any worksheet.")

    def _validate_required_columns(self) -> None:
        self.__say("Validating required columns [RC]:")

        for sheet_key, required_cols in DD_COLUMN_MAP.items():
            table_name = DD_TABLE_MAP.get(sheet_key)
            if not table_name or required_cols is None:
                self.__say(f"[RC] {sheet_key}: no table requirements")
                continue

            tkey = table_name.lower()
            tbl = self.tables.get(tkey)
            if tbl is None:
                raise DataDictionaryImportError(
                    f"Required table '{table_name}' for sheet '{sheet_key}' not loaded."
                )

            colmap = _table_column_index_map(tbl)
            self.table_column_maps[tkey] = colmap

            missing: list[str] = []
            for rc in required_cols:
                rc_key = (rc or "").strip().lower()
                if rc_key not in colmap:
                    missing.append(rc)

            if missing:
                raise DataDictionaryImportError(
                    f"Missing required columns in table '{table_name}' (sheet '{sheet_key}'): {', '.join(missing)}"
                )

            self.__say(
                f"[RC] {sheet_key}.{table_name}: OK ({len(required_cols)}/{len(required_cols)} present)"
            )

        self.__say("Required columns validated")


def export_dictionary(
    dictionary: Dictionary,
    directory: Path | str,
    filename: str | None = None,
    overwrite: bool = False,
    debug: bool = False,
    _template_path: Path | str | None = None,
) -> None:
    """
    Summary:
        Export a data dictionary to an Excel file.

    Arguments:
        dictionary (Dictionary): data dictionary to export
        directory (Path | str): directory to export to
        filename (str | None): filename to export to (default is None)
        overwrite (bool): whether to overwrite existing file (default is False)
        debug (bool): whether to print debug information (default is False)
        _template_path (Path | str | None): path to template data dictionary (default is None; changing not advised)

    Raises:
        DataDictionaryExportError: if unable to export data dictionary
    """
    exporter = Exporter(dictionary, debug=debug, template_path=_template_path)
    exporter.write_details()
    exporter.write_tables()
    exporter.write_columns()
    exporter.write_enumerations()
    exporter.export(directory=directory, filename=filename, overwrite=overwrite)
