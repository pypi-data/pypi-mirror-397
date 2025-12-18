from copy import copy
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Mapping, Tuple

from openpyxl.formatting.formatting import (
    ConditionalFormatting,
    ConditionalFormattingList,
)
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.worksheet import Worksheet

from valediction.exceptions import DataDictionaryError, DataDictionaryImportError


@dataclass
class CFRuleInfo:
    ranges: list[str]
    type: str | None = None
    formula: list[str] | None = None
    operator: str | None = None
    dxfId: int | None = None
    priority: int | None = None


@dataclass
class DVRuleInfo:
    ranges: list[str]
    type: str | None = None
    operator: str | None = None
    formula1: str | None = None
    formula2: str | None = None
    allowBlank: bool | None = None
    showErrorMessage: bool | None = None
    showInputMessage: bool | None = None
    promptTitle: str | None = None
    prompt: str | None = None
    errorTitle: str | None = None
    error: str | None = None


@dataclass
class CalculatedColInfo:
    column_name: str
    column_index: int
    header_cell: str
    formula_sample: str
    coverage: float


def _collect_conditional_formats(worksheet: Worksheet) -> Iterable[CFRuleInfo]:
    """Yield CF rules for a worksheet with their target ranges.

    Works across openpyxl versions by checking internal/public mappings.
    """
    cf = getattr(worksheet, "conditional_formatting", None)
    if cf is None:
        return
        yield  # for typing

    items = None
    # Preferred: internal mapping {sqref(tuple-of-ranges or string): [Rule,...]}
    if hasattr(cf, "_cf_rules") and isinstance(cf._cf_rules, dict):
        items = list(cf._cf_rules.items())
    # Fallback for some versions
    elif hasattr(cf, "cf_rules") and hasattr(cf, "ranges"):
        items = [(rng, cf.cf_rules(rng)) for rng in cf.ranges]

    if not items:
        return

    for sqref, rules in items:
        # normalize ranges to list[str]
        if isinstance(sqref, (list, tuple)):
            ranges = [str(r) for r in sqref]
        else:
            ranges = [str(sqref)]
        for rule in rules:
            yield CFRuleInfo(
                ranges=ranges,
                type=getattr(rule, "type", None),
                formula=getattr(rule, "formula", None),
                operator=getattr(rule, "operator", None),
                dxfId=getattr(rule, "dxfId", None),
                priority=getattr(rule, "priority", None),
            )


def _collect_data_validations(worksheet: Worksheet) -> Iterable[DVRuleInfo]:
    """Yield data validation rules for a worksheet with their target ranges."""
    dv_list = getattr(worksheet, "data_validations", None)
    if dv_list is None:
        return
        yield  # typing

    # In openpyxl, ws.data_validations.dataValidation is a list of DataValidation
    dvals = getattr(dv_list, "dataValidation", None)
    if not dvals:
        return

    for dv in dvals:
        # sqref is a MultiCellRange; convert to list[str]
        ranges: list[str] = []
        sqref = getattr(dv, "sqref", None) or getattr(dv, "ranges", None)
        if sqref is not None:
            try:
                # MultiCellRange: iterate .ranges -> CellRange objects
                ranges = [str(r) for r in sqref.ranges]
            except Exception:
                # fallback to string
                ranges = [str(sqref)]

        yield DVRuleInfo(
            ranges=ranges,
            type=getattr(dv, "type", None),
            operator=getattr(dv, "operator", None),
            formula1=getattr(dv, "formula1", None),
            formula2=getattr(dv, "formula2", None),
            allowBlank=getattr(dv, "allowBlank", None),
            showErrorMessage=getattr(dv, "showErrorMessage", None),
            showInputMessage=getattr(dv, "showInputMessage", None),
            promptTitle=getattr(dv, "promptTitle", None),
            prompt=getattr(dv, "prompt", None),
            errorTitle=getattr(dv, "errorTitle", None),
            error=getattr(dv, "error", None),
        )


def _collect_table_formulas(table: ExcelTable) -> dict[str, str]:
    out: dict[str, str] = {}
    for column in getattr(table, "tableColumns", []):
        name = (getattr(column, "name", "") or "").strip()
        if not name:
            continue
        formula = getattr(column, "calculatedColumnFormula", None)
        if not formula:
            continue
        text = (
            getattr(formula, "attr_text", None)
            or getattr(formula, "text", None)
            or str(formula)
        )
        if text and not str(text).startswith("="):
            text = f"={text}"
        out[name.lower()] = str(text)
    return out


def _table_column_index_map(tbl: ExcelTable) -> dict[str, int]:
    cols = getattr(tbl, "tableColumns", None)
    if not cols:
        return {}

    name_to_idx: dict[str, int] = {}
    for idx, col in enumerate(list(cols), start=1):
        raw = (getattr(col, "name", "") or "").strip()
        if not raw:
            continue
        key = raw.lower()
        if key in name_to_idx:
            raise DataDictionaryImportError(
                f"Duplicate column name (case-insensitive) in table '{getattr(tbl, 'displayName', getattr(tbl, 'name', ''))}': '{raw}'"
            )
        name_to_idx[key] = idx
    return name_to_idx


def _cf_entries(cf: ConditionalFormattingList) -> list[tuple[object, list]]:
    """Return [(key, rules)] from ws.conditional_formatting or [] if unsupported."""
    if hasattr(cf, "_cf_rules") and isinstance(cf._cf_rules, dict):
        return list(cf._cf_rules.items())
    return []


def _key_ranges(key: ConditionalFormatting | MultiCellRange) -> list[str]:
    """Normalize a CF dict key into a list of A1 range strings."""
    if hasattr(key, "sqref") and hasattr(key.sqref, "ranges"):  # ConditionalFormatting
        return [str(r) for r in key.sqref.ranges]
    if hasattr(key, "ranges"):  # MultiCellRange
        return [str(r) for r in key.ranges]
    if isinstance(key, (list, tuple)):
        return [str(r) for r in key]
    return [str(key)]


def _extend_ranges(
    range_strs: List[str], t_min_c: int, t_max_c: int, old_bottom: int, new_row: int
) -> Tuple[List[str], bool]:
    """Compute new sqref ranges; return (ranges, changed?)."""
    new_ranges: List[str] = []
    changed = False
    for rng in range_strs:
        try:
            c1, r1, c2, r2 = range_boundaries(rng)
        except ValueError:
            new_ranges.append(rng)
            continue
        if r1 <= new_row <= r2:
            new_ranges.append(rng)
            continue
        overlaps_cols = not (c2 < t_min_c or c1 > t_max_c)
        if r2 == old_bottom and overlaps_cols:
            r2 = new_row
            changed = True
        new_ranges.append(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")
    return new_ranges, changed


def _extend_cf_for_new_row(
    ws, table_cols: tuple[int, int], old_bottom: int, new_row: int
) -> None:
    cf = getattr(ws, "conditional_formatting", None)
    if cf is None:
        return

    t_min_c, t_max_c = table_cols
    entries = _cf_entries(cf)
    if not entries:
        return

    for key, rules in entries:
        range_strs = _key_ranges(key)
        new_ranges, changed = _extend_ranges(
            range_strs, t_min_c, t_max_c, old_bottom, new_row
        )
        if not changed:
            continue
        # replace mapping (do NOT mutate key)
        cf._cf_rules.pop(key, None)
        sqref = ",".join(new_ranges)
        for rule in rules:
            cf.add(sqref, copy(rule))


def _extend_dv_for_new_row(ws, table_cols: tuple[int, int], new_row: int) -> None:
    """For each DataValidation object, if any of its ranges overlap table columns, add
    the new_row segment for the overlapping columns to that DV."""
    t_min_c, t_max_c = table_cols
    dv_list = getattr(ws, "data_validations", None)
    if not dv_list:
        return

    dvals = getattr(dv_list, "dataValidation", None)
    if not dvals:
        return

    for dv in dvals:
        sqref = getattr(dv, "sqref", None) or getattr(dv, "ranges", None)
        if not sqref:
            continue
        try:
            rngs = list(sqref.ranges)  # MultiCellRange -> list[CellRange]
        except Exception:
            rngs = []

        to_add_segments: list[str] = []
        for cr in rngs:
            c1, r1, c2, r2 = range_boundaries(str(cr))
            # If DV already covers the new row in this segment, skip
            if r1 <= new_row <= r2:
                continue
            # Column overlap with the table
            oc1 = max(c1, t_min_c)
            oc2 = min(c2, t_max_c)
            if oc1 > oc2:
                continue
            to_add_segments.append(
                f"{get_column_letter(oc1)}{new_row}:{get_column_letter(oc2)}{new_row}"
            )

        for seg in to_add_segments:
            dv.add(seg)


def _first_blank_data_row(
    ws: Worksheet, tbl: ExcelTable, colmap: dict[str, int], formulas: dict[str, str]
) -> int | None:
    """Return the first data-row index (0-based worksheet row) that is blank across all
    *non-formula* columns.

    None if no such row.
    """
    min_c, min_r, max_c, max_r = range_boundaries(tbl.ref)
    data_start = min_r + 1
    if data_start > max_r:
        return None

    # quick reverse map: table-index(1..n) -> header_lower
    idx_to_header = {v: k for k, v in colmap.items()}
    ncols = max_c - min_c + 1

    for r in range(data_start, max_r + 1):
        all_blank = True
        for j_idx in range(1, ncols + 1):
            header_lower = idx_to_header.get(j_idx, "")
            if header_lower in formulas:
                # ignore formula columns when deciding "blank"
                continue
            cell = ws.cell(r, min_c + j_idx - 1)
            val = cell.value
            if val is None:
                continue
            if isinstance(val, str) and val.strip() == "":
                continue
            # any non-empty value in a non-formula column -> row not blank
            all_blank = False
            break
        if all_blank:
            return r
    return None


def _build_row_list_from_mapping(
    tbl: ExcelTable,
    colmap: Dict[str, int],
    formulas: Dict[str, str],
    mapping: Mapping[str, Any],
) -> List[Any]:
    """Build a full-width list for add_row() from a {header->value} mapping (CI).

    Values for formula columns are ignored by add_row anyway; we leave them None.
    """
    min_c, min_r, max_c, max_r = range_boundaries(tbl.ref)
    ncols = max_c - min_c + 1

    # normalise input keys to lower-case once
    src = {str(k).strip().lower(): v for k, v in mapping.items()}
    # start as all None
    row_list: List[Any] = [None] * ncols

    # inverse map: table index -> header lower
    idx_to_header = {v: k for k, v in colmap.items()}

    for j_idx in range(1, ncols + 1):
        header_lower = idx_to_header.get(j_idx, "")
        if not header_lower:
            continue
        # skip formula columns; add_row will write formulas
        if header_lower in formulas:
            continue
        if header_lower in src:
            row_list[j_idx - 1] = src[header_lower]
    return row_list


def _norm_label(s: object) -> str:
    """Normalise a cell label for case-insensitive comparison."""
    if s is None:
        return ""
    txt = str(s).replace("\n", " ").replace("\r", " ").strip()
    # ignore a single trailing colon, collapse inner whitespace
    if txt.endswith(":"):
        txt = txt[:-1]
    return " ".join(txt.split()).casefold()


def _find_label_cell(ws: Worksheet, label: str) -> tuple[int, int]:
    """Find the cell coordinates (row, col) whose text matches `label` case-
    insensitively (ignoring a trailing colon).

    Scans the used range only. Raises DataDictionaryError if not found.
    """
    target = _norm_label(label)
    min_row, min_col = ws.min_row, ws.min_column
    max_row, max_col = ws.max_row, ws.max_column
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            if _norm_label(ws.cell(r, c).value) == target:
                return r, c
    raise DataDictionaryError(f"Details label not found: {label!r}")
